"""End-to-end two-player game over WebSockets."""

import io

import pytest
from conftest import SUBJECT, make_question, make_test_payload, signup

pytestmark = pytest.mark.filterwarnings("ignore::DeprecationWarning")


def wait_for(ws, msg_type, limit=60):
    """Drain messages until one of `msg_type` arrives, returning it."""
    seen = []
    for _ in range(limit):
        msg = ws.receive_json()
        seen.append(msg.get("type"))
        if msg.get("type") == msg_type:
            return msg
    raise AssertionError(f"never received {msg_type!r}; saw {seen}")


@pytest.fixture()
def authed(client):
    signup(client)
    return client


@pytest.fixture()
def test_id(authed):
    payload = make_test_payload(
        questions=[
            make_question(1, q="Q1 SI unit of force?", correct=0),
            make_question(2, q="Q2 SI unit of power?", options=["Newton", "Joule", "Watt", "Pascal"], correct=2),
        ]
    )
    return authed.post(f"/api/tests/{SUBJECT}", json=payload).json()["test"]["id"]


def test_two_player_game_end_to_end(authed, test_id):
    with authed.websocket_connect("/ws?visitorId=host-1") as host:
        host.send_json({"action": "host_join", "subject": SUBJECT, "testId": test_id, "sessionName": "Tutorial 3"})
        joined = wait_for(host, "host_joined")
        assert joined["totalQ"] == 2
        assert joined["hasQuestions"] is True
        assert joined["selectedTest"]["title"] == "Chapter 1 Quiz"

        with authed.websocket_connect("/ws?visitorId=player-a") as pa, \
             authed.websocket_connect("/ws?visitorId=player-b") as pb:
            pa.send_json({"action": "player_join", "name": "Ada", "studentNumber": "221000001", "subject": SUBJECT})
            assert wait_for(pa, "joined")["playerId"] == "player-a"

            pb.send_json({"action": "player_join", "name": "Ben", "studentNumber": "221000002", "subject": SUBJECT})
            assert wait_for(pb, "joined")["playerId"] == "player-b"

            update = wait_for(host, "player_update")
            assert sorted(p["name"] for p in update["players"]) == ["Ada", "Ben"]

            host.send_json({"action": "start_game", "shuffle": False, "useCode": False})

            # ── Question 1 ────────────────────────────────────────────────────
            wait_for(pa, "get_ready")
            wait_for(pb, "get_ready")
            q1 = wait_for(pa, "question")
            assert q1["qNum"] == 1 and q1["totalQ"] == 2
            assert q1["timeLimit"] == 30
            wait_for(pb, "question")
            host_q1 = wait_for(host, "question")
            assert host_q1["correctAnswer"] == 0

            pa.send_json({"action": "answer", "choice": 0})   # correct
            result_a = wait_for(pa, "answer_result")
            assert result_a["correct"] is True
            assert result_a["points"] > 0

            pb.send_json({"action": "answer", "choice": 1})   # wrong
            result_b = wait_for(pb, "answer_result")
            assert result_b["correct"] is False
            assert result_b["points"] == 0

            # Everyone answered → the question finishes early without waiting
            # out the 30 second timer.
            reveal = wait_for(host, "reveal")
            assert reveal["correctAnswer"] == 0
            assert reveal["isLast"] is False
            lb = wait_for(pa, "leaderboard")["leaderboard"]
            assert lb[0]["name"] == "Ada"

            # ── Question 2 (auto-advanced after the reveal pause) ─────────────
            q2 = wait_for(pa, "question")
            assert q2["qNum"] == 2
            wait_for(pb, "question")

            pa.send_json({"action": "answer", "choice": 3})   # wrong
            wait_for(pa, "answer_result")
            pb.send_json({"action": "answer", "choice": 2})   # correct
            wait_for(pb, "answer_result")

            final_host = wait_for(host, "final")
            assert final_host["hasStats"] is True
            board = final_host["leaderboard"]
            assert {p["name"] for p in board} == {"Ada", "Ben"}
            assert [p["rank"] for p in board] == [1, 2]

            final_a = wait_for(pa, "final")
            assert len(final_a["leaderboard"]) == 2

    # ── Stats spreadsheet ────────────────────────────────────────────────────
    stats = authed.get(f"/api/stats/{SUBJECT}")
    assert stats.status_code == 200, stats.text
    assert "spreadsheetml" in stats.headers["content-type"]
    assert "Tutorial_3" in stats.headers["content-disposition"]

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(stats.content))
    rows = list(wb["Student Results"].iter_rows(values_only=True))
    assert rows[0][:4] == ("Rank", "Student Name", "Student Number", "Total Score")
    names = {row[1] for row in rows[1:]}
    assert names == {"Ada", "Ben"}
    numbers = {row[2] for row in rows[1:]}
    assert numbers == {"221000001", "221000002"}


def test_host_join_requires_lecturer_auth(client, server_module):
    signup(client)
    tid = client.post(f"/api/tests/{SUBJECT}", json=make_test_payload()).json()["test"]["id"]
    client.cookies.clear()
    with client.websocket_connect("/ws?visitorId=anon-host") as ws:
        ws.send_json({"action": "host_join", "subject": SUBJECT, "testId": tid})
        assert wait_for(ws, "auth_required")


def test_player_cannot_join_a_game_in_progress(authed, test_id):
    with authed.websocket_connect("/ws?visitorId=host-2") as host:
        host.send_json({"action": "host_join", "subject": SUBJECT, "testId": test_id})
        wait_for(host, "host_joined")

        with authed.websocket_connect("/ws?visitorId=early-bird") as pa:
            pa.send_json({"action": "player_join", "name": "Ada", "studentNumber": "1", "subject": SUBJECT})
            wait_for(pa, "joined")
            host.send_json({"action": "start_game", "shuffle": False, "useCode": False})
            wait_for(pa, "question")

            with authed.websocket_connect("/ws?visitorId=latecomer") as pc:
                pc.send_json({"action": "player_join", "name": "Cid", "studentNumber": "3", "subject": SUBJECT})
                err = wait_for(pc, "error")
                assert "already in progress" in err["message"]


def test_stats_download_404s_before_any_game(authed):
    assert authed.get(f"/api/stats/{SUBJECT}").status_code == 404
