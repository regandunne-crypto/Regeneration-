#!/usr/bin/env python3
"""WebSocket game server for multi-subject quiz platform with optional Supabase-backed test bank.

Key ideas:
- Students still join by subject.
- Host now selects a subject and then a saved test for that subject.
- Tests can be stored durably in Supabase when environment variables are configured.
- Without Supabase, the app still works using in-memory fallback storage (not durable).
"""

import asyncio
import base64
import hashlib
import hmac
import io
import json
import os
import random
import re
import secrets
import string
import time
import uuid
from contextlib import asynccontextmanager
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import httpx
from fastapi import FastAPI, File, HTTPException, Request, Response, UploadFile, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field, ValidationError, field_validator
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.errors import RateLimitExceeded
from slowapi.util import get_remote_address

import docx_import

# `datetime.UTC` only exists on Python 3.11+. Render pins a Python version per
# service, so importing it directly risks a boot failure on an older runtime.
# `timezone.utc` is identical and available everywhere.
UTC = timezone.utc

# ──────────────────────────────────────────────────────────────────────────────
# Subject catalogue and built-in legacy question sets
# ──────────────────────────────────────────────────────────────────────────────
SUBJECTS = {
    '1EM105B': {'code': '1EM105B', 'name': 'Mechanics', 'questions': []},
    'DYN317B': {'code': 'DYN317B', 'name': 'Dynamics', 'questions': []},
    'MEC105B': {'code': 'MEC105B', 'name': 'Mechanics', 'questions': []},
}

BUILTIN_SUBJECT_CODES = set(SUBJECTS.keys())
SUBJECT_CODE_PATTERN = re.compile(r"^[A-Z0-9]{3,10}$")

TIME_PER_Q = int(os.environ.get("TIME_PER_Q", "30"))
MAX_POINTS = 1000
MIN_POINTS = 200
# Bounds for a configurable question time limit. TIME_PER_Q above is now the
# fallback default rather than a hard constant.
MIN_TIME_LIMIT = 5
MAX_TIME_LIMIT = 300
REQUEST_TIMEOUT = 20

# Pacing constants. Overridable by environment so the test suite can run a whole
# game in well under a second instead of waiting out the real classroom timings.
GET_READY_SECONDS = float(os.environ.get("GET_READY_SECONDS", "3"))
REVEAL_SECONDS = float(os.environ.get("REVEAL_SECONDS", "5"))
GAME_CODE_COUNTDOWN_SECONDS = float(os.environ.get("GAME_CODE_COUNTDOWN_SECONDS", "20"))
REQUIRE_SUPABASE = os.environ.get("REQUIRE_SUPABASE", "").strip().lower() in {"1", "true", "yes", "on"}
_local_store_env = os.environ.get("LOCAL_STORE_PATH", "").strip()
LOCAL_STORE_PATH = Path(_local_store_env).expanduser() if _local_store_env else (Path(__file__).resolve().parent / "local_store.json")
LOCAL_STORE_VERSION = 1

limiter = Limiter(key_func=get_remote_address)

# Draft autosave fires 1.5 s after typing stops, so a lecturer writing a long
# quiz legitimately produces many writes a minute. A tight limit here would
# break the very feature Phase 2 fixed, so this is deliberately generous.
DRAFT_RATE_LIMIT = os.environ.get("DRAFT_RATE_LIMIT", "120/minute")

# Stored game results. The automatic spreadsheet download at the end of a game
# stays the authoritative record; these rows exist so a Render redeploy does not
# destroy results that were never downloaded. Off with one variable if the free
# plan's 500 MB ever feels tight.
PERSIST_RESULTS = os.environ.get("PERSIST_RESULTS", "true").strip().lower() not in {"0", "false", "no", "off"}
RESULTS_RETENTION = max(1, int(os.environ.get("RESULTS_RETENTION", "20")))


class SupabaseUnavailable(RuntimeError):
    pass


# ── PostgREST filter safety ──────────────────────────────────────────────────
# Values go straight into filter strings such as f"eq.{test_id}". PostgREST
# treats commas, dots and parentheses as filter syntax, so an unvalidated value
# can change which rows a query matches. Validate before building any query.

UUID_PATTERN = re.compile(r"^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}$")
LOCAL_ID_PATTERN = re.compile(r"^local:[0-9a-fA-F-]{36}$")


def safe_test_id(test_id: str) -> str:
    value = (test_id or "").strip()
    if UUID_PATTERN.match(value) or LOCAL_ID_PATTERN.match(value):
        return value
    raise ValueError("Invalid test id.")


def safe_subject_code(subject_code: str) -> str:
    value = (subject_code or "").strip().upper()
    if SUBJECT_CODE_PATTERN.match(value):
        return value
    raise ValueError("Invalid subject code.")


def safe_lecturer_id(lecturer_id: str) -> str:
    value = (lecturer_id or "").strip()
    if UUID_PATTERN.match(value):
        return value
    raise ValueError("Invalid lecturer id.")


# ──────────────────────────────────────────────────────────────────────────────
# Test bank models + storage
# ──────────────────────────────────────────────────────────────────────────────
SESSION_COOKIE_NAME = "lecturer_session"
SESSION_MAX_AGE = 60 * 60 * 24 * 14  # 14 days


DEV_SESSION_SECRET = "engineering-quiz-dev-secret"


def _session_secret() -> str:
    return (
        os.environ.get("APP_SESSION_SECRET", "").strip()
        or os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "").strip()
        or DEV_SESSION_SECRET
    )


def _check_session_secret_configured() -> None:
    """Refuse to start in production with the built-in development secret.

    Anyone who has read this source could otherwise forge a lecturer session
    cookie. Locally (no SUPABASE_URL) the fallback is fine and keeps first-run
    setup frictionless.

    NOTE for the first deploy: set APP_SESSION_SECRET to the *current* value of
    SUPABASE_SERVICE_ROLE_KEY. That is what existing sessions are already signed
    with, so nobody gets signed out. Rotate it to a fresh random value later.
    """
    if not os.environ.get("SUPABASE_URL", "").strip():
        return
    if os.environ.get("APP_SESSION_SECRET", "").strip():
        return
    if os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "").strip():
        # Sessions are signed with the service role key. Works, but the key is
        # meant to be rotatable without signing every lecturer out.
        print(
            "[startup] WARNING: APP_SESSION_SECRET is not set, so lecturer sessions are "
            "signed with SUPABASE_SERVICE_ROLE_KEY. Set APP_SESSION_SECRET to the current "
            "value of SUPABASE_SERVICE_ROLE_KEY now (nobody gets signed out), then rotate "
            "it to a fresh random value at a convenient moment."
        )
        return
    raise RuntimeError(
        "APP_SESSION_SECRET must be set when SUPABASE_URL is configured. "
        "Refusing to sign lecturer sessions with the built-in development secret."
    )


# ── Lecturer signup gating ───────────────────────────────────────────────────
# If neither variable is set, signup stays open exactly as before. This matters:
# the change must not lock anyone out of a running deployment before the owner
# has configured it.

def _allowed_email_domains() -> list[str]:
    raw = os.environ.get("ALLOWED_EMAIL_DOMAINS", "").strip()
    return [d.strip().lower().lstrip("@") for d in raw.split(",") if d.strip()]


def _signup_invite_code() -> str:
    return os.environ.get("SIGNUP_INVITE_CODE", "").strip()


def check_signup_allowed(email: str, invite_code: str) -> None:
    """Raise PermissionError with a message the lecturer can act on."""
    domains = _allowed_email_domains()
    required_code = _signup_invite_code()
    if not domains and not required_code:
        return                                  # open signup, as before

    domain = (email or "").rsplit("@", 1)[-1].strip().lower()
    domain_ok = bool(domains) and any(domain == d or domain.endswith("." + d) for d in domains)
    code_ok = bool(required_code) and hmac.compare_digest((invite_code or "").strip(), required_code)

    if domain_ok or code_ok:
        return

    if domains and required_code:
        listed = ", ".join(domains)
        raise PermissionError(
            f"Accounts are limited to {listed} email addresses. Use your work email, "
            "or enter the invite code your administrator gave you."
        )
    if domains:
        listed = ", ".join(domains)
        raise PermissionError(
            f"Accounts are limited to {listed} email addresses. Please sign up with your work email."
        )
    raise PermissionError("An invite code is required to create a lecturer account.")


def hash_password(password: str) -> str:
    salt = secrets.token_hex(16)
    rounds = 260_000
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), bytes.fromhex(salt), rounds)
    return f"pbkdf2_sha256${rounds}${salt}${digest.hex()}"


def verify_password(password: str, stored: str) -> bool:
    try:
        algorithm, rounds_str, salt, digest = stored.split("$", 3)
        if algorithm != "pbkdf2_sha256":
            return False
        rounds = int(rounds_str)
    except Exception:
        return False
    candidate = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), bytes.fromhex(salt), rounds).hex()
    return hmac.compare_digest(candidate, digest)


# 260k PBKDF2 rounds take 100-200 ms. Run them on a worker thread: on the event
# loop they stall every other request, including live games — a class signing in
# would freeze the quiz everyone else is playing.
async def hash_password_async(password: str) -> str:
    return await asyncio.to_thread(hash_password, password)


async def verify_password_async(password: str, stored: str) -> bool:
    return await asyncio.to_thread(verify_password, password, stored)


def create_session_token(lecturer_id: str) -> str:
    expires = int(time.time()) + SESSION_MAX_AGE
    payload = f"{lecturer_id}.{expires}"
    signature = hmac.new(_session_secret().encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).digest()
    return base64.urlsafe_b64encode(f"{payload}.{signature.hex()}".encode()).decode("utf-8")


def parse_session_token(token: str | None) -> str | None:
    if not token:
        return None
    try:
        decoded = base64.urlsafe_b64decode(token.encode("utf-8")).decode("utf-8")
        lecturer_id, expires_str, signature = decoded.split(".", 2)
        payload = f"{lecturer_id}.{expires_str}"
        expected = hmac.new(_session_secret().encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(signature, expected):
            return None
        if int(expires_str) < int(time.time()):
            return None
        return lecturer_id
    except Exception:
        return None


# ── Visitor identity ─────────────────────────────────────────────────────────
# The visitor id used to come straight off a query parameter, so a student could
# supply someone else's id and take over their session and score. The server now
# issues the id inside a signed token and never trusts a bare id.

VISITOR_TOKEN_MAX_AGE = 60 * 60 * 24 * 30      # 30 days


def create_visitor_token(visitor_id: str | None = None) -> tuple[str, str]:
    visitor_id = visitor_id or str(uuid.uuid4())
    expires = int(time.time()) + VISITOR_TOKEN_MAX_AGE
    payload = f"{visitor_id}.{expires}"
    signature = hmac.new(_session_secret().encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()
    token = base64.urlsafe_b64encode(f"{payload}.{signature}".encode()).decode("utf-8")
    return visitor_id, token


def parse_visitor_token(token: str | None) -> str | None:
    if not token:
        return None
    try:
        decoded = base64.urlsafe_b64decode(token.encode("utf-8")).decode("utf-8")
        visitor_id, expires_str, signature = decoded.rsplit(".", 2)
        payload = f"{visitor_id}.{expires_str}"
        expected = hmac.new(_session_secret().encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(signature, expected):
            return None
        if int(expires_str) < int(time.time()):
            return None
        return visitor_id or None
    except Exception:
        return None


class QuestionPayload(BaseModel):
    q: str = Field(min_length=1, max_length=600)
    options: list[str]
    correct: int = Field(ge=0, le=3)
    explanation: str = Field(default="", max_length=2000)
    # None means "use the test default". Existing saved questions have no such
    # field at all, so they keep running at exactly TIME_PER_Q with no migration.
    time_limit: int | None = Field(default=None, ge=MIN_TIME_LIMIT, le=MAX_TIME_LIMIT)

    @field_validator("options")
    @classmethod
    def validate_options(cls, value: list[str]) -> list[str]:
        if len(value) != 4:
            raise ValueError("Each question must have exactly 4 options.")
        cleaned = []
        for item in value:
            text = (item or "").strip()
            if not text:
                raise ValueError("Answer options cannot be blank.")
            cleaned.append(text[:240])
        return cleaned

    @field_validator("q")
    @classmethod
    def validate_question_text(cls, value: str) -> str:
        value = value.strip()
        if not value:
            raise ValueError("Question text cannot be blank.")
        return value

    @field_validator("explanation")
    @classmethod
    def normalize_explanation(cls, value: str) -> str:
        return (value or "").strip()


class TestPayload(BaseModel):
    title: str = Field(min_length=1, max_length=140)
    chapter: str = Field(default="", max_length=140)
    description: str = Field(default="", max_length=600)
    questions: list[QuestionPayload]
    default_time_limit: int = Field(default=TIME_PER_Q, ge=MIN_TIME_LIMIT, le=MAX_TIME_LIMIT)

    @field_validator("title")
    @classmethod
    def validate_title(cls, value: str) -> str:
        value = value.strip()
        if not value:
            raise ValueError("Test title cannot be blank.")
        return value

    @field_validator("chapter")
    @classmethod
    def normalize_chapter(cls, value: str) -> str:
        return (value or "").strip()

    @field_validator("description")
    @classmethod
    def normalize_description(cls, value: str) -> str:
        return (value or "").strip()

    @field_validator("questions")
    @classmethod
    def validate_questions(cls, value: list[QuestionPayload]) -> list[QuestionPayload]:
        if not value:
            raise ValueError("A test must include at least one question.")
        return value


class SubjectPayload(BaseModel):
    code: str = Field(min_length=3, max_length=10)
    name: str = Field(min_length=2, max_length=60)

    @field_validator("code")
    @classmethod
    def normalize_code(cls, value: str) -> str:
        cleaned = (value or "").strip().upper()
        if not SUBJECT_CODE_PATTERN.match(cleaned):
            raise ValueError("Subject code must be 3-10 letters or numbers (no spaces).")
        return cleaned

    @field_validator("name")
    @classmethod
    def normalize_name(cls, value: str) -> str:
        cleaned = (value or "").strip()
        if len(cleaned) < 2 or len(cleaned) > 60:
            raise ValueError("Subject name must be 2-60 characters.")
        return cleaned


class LecturerSignupPayload(BaseModel):
    name: str = Field(min_length=2, max_length=120)
    email: str = Field(min_length=5, max_length=240)
    password: str = Field(min_length=8, max_length=200)
    inviteCode: str = Field(default="", max_length=200)

    @field_validator("name")
    @classmethod
    def validate_name(cls, value: str) -> str:
        value = (value or "").strip()
        if len(value) < 2:
            raise ValueError("Name must be at least 2 characters.")
        return value

    @field_validator("email")
    @classmethod
    def validate_email(cls, value: str) -> str:
        value = (value or "").strip().lower()
        if "@" not in value or "." not in value.split("@")[-1]:
            raise ValueError("Please enter a valid email address.")
        return value


class LecturerLoginPayload(BaseModel):
    email: str = Field(min_length=5, max_length=240)
    password: str = Field(min_length=8, max_length=200)

    @field_validator("email")
    @classmethod
    def normalize_email(cls, value: str) -> str:
        return (value or "").strip().lower()


class DraftQuestionPayload(BaseModel):
    q: str = Field(default="", max_length=600)
    options: list[str] = Field(default_factory=lambda: ["", "", "", ""])
    correct: int = Field(default=0, ge=0, le=3)
    explanation: str = Field(default="", max_length=2000)
    time_limit: int | None = Field(default=None, ge=MIN_TIME_LIMIT, le=MAX_TIME_LIMIT)

    @field_validator("options")
    @classmethod
    def validate_options(cls, value: list[str]) -> list[str]:
        value = list(value or [])[:4]
        while len(value) < 4:
            value.append("")
        return [(item or "").strip()[:240] for item in value]

    @field_validator("q")
    @classmethod
    def normalize_question_text(cls, value: str) -> str:
        return (value or "").strip()

    @field_validator("explanation")
    @classmethod
    def normalize_explanation(cls, value: str) -> str:
        return (value or "").strip()


class DraftPayload(BaseModel):
    title: str = Field(default="", max_length=140)
    chapter: str = Field(default="", max_length=140)
    description: str = Field(default="", max_length=600)
    questions: list[DraftQuestionPayload] = Field(default_factory=list)
    editingTestId: str | None = None
    default_time_limit: int = Field(default=TIME_PER_Q, ge=MIN_TIME_LIMIT, le=MAX_TIME_LIMIT)

    @field_validator("title")
    @classmethod
    def normalize_title(cls, value: str) -> str:
        return (value or "").strip()

    @field_validator("chapter")
    @classmethod
    def normalize_chapter(cls, value: str) -> str:
        return (value or "").strip()

    @field_validator("description")
    @classmethod
    def normalize_description(cls, value: str) -> str:
        return (value or "").strip()


class SupabaseStore:
    def __init__(self, base_url: str, service_role_key: str):
        self.base_url = base_url.rstrip("/")
        self.service_role_key = service_role_key
        self._client = httpx.AsyncClient(
            headers={
                "apikey": self.service_role_key,
                "Authorization": f"Bearer {self.service_role_key}",
                "Content-Type": "application/json",
            },
            timeout=REQUEST_TIMEOUT,
        )
        self.quiz_tests_base = f"{self.base_url}/rest/v1/quiz_tests"
        self.lecturers_base = f"{self.base_url}/rest/v1/quiz_lecturers"
        self.drafts_base = f"{self.base_url}/rest/v1/quiz_test_drafts"
        self.subjects_base = f"{self.base_url}/rest/v1/quiz_subjects"
        self.results_base = f"{self.base_url}/rest/v1/quiz_game_results"

    async def aclose(self) -> None:
        await self._client.aclose()

    def _check_response(self, resp: httpx.Response) -> None:
        if not resp.is_success:
            try:
                detail = resp.json()
            except Exception:
                detail = resp.text
            raise RuntimeError(f"Supabase request failed: {detail}")

    # Characters PostgREST reads as filter syntax. A value containing these can
    # change which rows a query matches, so no interpolated value may carry them.
    _FILTER_METACHARS = set(',()"*')
    _FILTER_OPERATORS = ("eq.", "ilike.", "like.", "neq.", "gt.", "gte.", "lt.", "lte.", "in.", "is.")

    @classmethod
    def _check_filter_params(cls, params: dict[str, Any] | None) -> None:
        for key, value in (params or {}).items():
            if key in {"select", "order", "limit", "offset"} or not isinstance(value, str):
                continue
            if not value.startswith(cls._FILTER_OPERATORS):
                continue
            operand = value.split(".", 1)[1]
            if cls._FILTER_METACHARS & set(operand):
                raise ValueError(f"Refusing to build a PostgREST filter from an unsafe value for {key!r}.")

    async def _request(self, method: str, url: str, *, params=None, body=None, prefer: str | None = None) -> list[dict[str, Any]]:
        # Single choke point: every filter value passes through here, so this
        # covers present and future call sites alike.
        self._check_filter_params(params)
        headers = dict(self._client.headers)
        if prefer:
            headers["Prefer"] = prefer
        resp = await self._client.request(method, url, params=params, headers=headers, json=body)
        self._check_response(resp)
        if not resp.text:
            return []
        try:
            return resp.json()
        except Exception:
            return []

    async def get_lecturer_by_email(self, email: str) -> dict[str, Any] | None:
        rows = await self._request("GET", self.lecturers_base, params={
            "select": "id,name,email,password_hash,created_at,updated_at",
            "email": f"eq.{email.lower()}",
            "limit": "1",
        })
        return rows[0] if rows else None

    async def get_lecturer_by_id(self, lecturer_id: str) -> dict[str, Any] | None:
        rows = await self._request("GET", self.lecturers_base, params={
            "select": "id,name,email,created_at,updated_at",
            "id": f"eq.{lecturer_id}",
            "limit": "1",
        })
        return rows[0] if rows else None

    async def create_lecturer(self, name: str, email: str, password_hash: str) -> dict[str, Any]:
        rows = await self._request("POST", self.lecturers_base, body={
            "name": name,
            "email": email.lower(),
            "password_hash": password_hash,
        }, prefer="return=representation")
        if not rows:
            raise RuntimeError("Supabase did not return the created lecturer.")
        return rows[0]

    async def list_subjects(self) -> list[dict[str, Any]]:
        rows = await self._request("GET", self.subjects_base, params={
            "select": "code,name,created_by,created_at",
            "order": "name.asc",
        })
        for row in rows:
            if row.get("code"):
                row["code"] = str(row["code"]).strip().upper()
            if row.get("name"):
                row["name"] = str(row["name"]).strip()
        return rows

    async def get_subject(self, code: str) -> dict[str, Any] | None:
        rows = await self._request("GET", self.subjects_base, params={
            "select": "code,name,created_by,created_at",
            "code": f"ilike.{code}",
            "limit": "1",
        })
        if not rows:
            return None
        row = rows[0]
        row["code"] = str(row.get("code") or "").strip().upper()
        row["name"] = str(row.get("name") or "").strip()
        return row

    async def create_subject(self, code: str, name: str, lecturer_id: str) -> dict[str, Any]:
        rows = await self._request("POST", self.subjects_base, body={
            "code": code,
            "name": name,
            "created_by": lecturer_id,
        }, prefer="return=representation")
        if not rows:
            raise RuntimeError("Supabase did not return the created subject.")
        row = rows[0]
        row["code"] = str(row.get("code") or "").strip().upper()
        row["name"] = str(row.get("name") or "").strip()
        return row

    async def delete_subject(self, code: str, lecturer_id: str) -> None:
        await self._request("DELETE", self.subjects_base, params={
            "code": f"ilike.{code}",
            "created_by": f"eq.{lecturer_id}",
        })

    async def subject_has_tests(self, subject_code: str) -> bool:
        rows = await self._request("GET", self.quiz_tests_base, params={
            "select": "id",
            "subject_code": f"eq.{subject_code}",
            "limit": "1",
        })
        return bool(rows)

    async def list_tests_by_creator(self, lecturer_id: str) -> list[dict[str, Any]]:
        rows = await self._request("GET", self.quiz_tests_base, params={
            "select": "id,subject_code,title,chapter,description,questions,question_count,default_time_limit,created_at,updated_at,created_by,owner_name",
            "created_by": f"eq.{lecturer_id}",
            "order": "subject_code.asc,updated_at.desc",
        })
        for row in rows:
            row["source"] = "supabase"
            row.setdefault("question_count", len(row.get("questions") or []))
        return rows

    async def list_all_test_counts(self) -> list[dict[str, Any]]:
        """Every test's subject and question count, for the subject list."""
        return await self._request("GET", self.quiz_tests_base, params={
            "select": "subject_code,question_count",
        })

    async def list_tests(self, subject_code: str, lecturer_id: str | None = None) -> list[dict[str, Any]]:
        rows = await self._request("GET", self.quiz_tests_base, params={
            "select": "id,subject_code,title,chapter,description,question_count,default_time_limit,created_at,updated_at,created_by,owner_name",
            "subject_code": f"eq.{subject_code}",
            "order": "updated_at.desc",
        })
        for row in rows:
            row["source"] = "supabase"
            row["can_edit"] = bool(lecturer_id and row.get("created_by") == lecturer_id)
        return rows

    async def get_test(self, subject_code: str, test_id: str, lecturer_id: str | None = None) -> dict[str, Any] | None:
        rows = await self._request("GET", self.quiz_tests_base, params={
            "select": "id,subject_code,title,chapter,description,questions,question_count,created_at,updated_at,created_by,owner_name",
            "subject_code": f"eq.{subject_code}",
            "id": f"eq.{test_id}",
            "limit": "1",
        })
        if not rows:
            return None
        row = rows[0]
        row["source"] = "supabase"
        row["can_edit"] = bool(lecturer_id and row.get("created_by") == lecturer_id)
        return row

    async def create_test(self, subject_code: str, payload: TestPayload, lecturer: dict[str, Any]) -> dict[str, Any]:
        rows = await self._request("POST", self.quiz_tests_base, body={
            "subject_code": subject_code,
            "title": payload.title,
            "chapter": payload.chapter or None,
            "description": payload.description or None,
            "question_count": len(payload.questions),
            "default_time_limit": payload.default_time_limit,
            "questions": [q.model_dump() for q in payload.questions],
            "created_by": lecturer["id"],
            "updated_by": lecturer["id"],
            "owner_name": lecturer.get("name") or lecturer.get("email") or "Lecturer",
        }, prefer="return=representation")
        if not rows:
            raise RuntimeError("Supabase did not return the created test.")
        row = rows[0]
        row["source"] = "supabase"
        row["can_edit"] = True
        return row

    async def delete_test(self, subject_code: str, test_id: str) -> None:
        await self._request("DELETE", self.quiz_tests_base, params={
            "subject_code": f"eq.{subject_code}",
            "id": f"eq.{test_id}",
        })

    async def update_test(self, subject_code: str, test_id: str, payload: TestPayload, lecturer: dict[str, Any]) -> dict[str, Any]:
        existing = await self.get_test(subject_code, test_id, lecturer["id"])
        if not existing:
            raise KeyError("Test not found")
        if existing.get("created_by") and existing.get("created_by") != lecturer["id"]:
            raise PermissionError("Only the lecturer who created this test can edit it.")
        rows = await self._request("PATCH", self.quiz_tests_base, params={
            "subject_code": f"eq.{subject_code}",
            "id": f"eq.{test_id}",
        }, body={
            "title": payload.title,
            "chapter": payload.chapter or None,
            "description": payload.description or None,
            "question_count": len(payload.questions),
            "default_time_limit": payload.default_time_limit,
            "questions": [q.model_dump() for q in payload.questions],
            "updated_by": lecturer["id"],
            "owner_name": existing.get("owner_name") or lecturer.get("name") or lecturer.get("email") or "Lecturer",
            "updated_at": datetime.now(UTC).isoformat(),
        }, prefer="return=representation")
        if not rows:
            raise RuntimeError("Supabase did not return the updated test.")
        row = rows[0]
        row["source"] = "supabase"
        row["can_edit"] = True
        return row

    async def list_game_results(self, subject_code: str, limit: int = 50) -> list[dict[str, Any]]:
        return await self._request("GET", self.results_base, params={
            "select": "id,subject_code,test_id,test_title,session_name,played_at,player_count,question_count",
            "subject_code": f"eq.{subject_code}",
            "order": "played_at.desc",
            "limit": str(limit),
        })

    async def get_game_result(self, result_id: str) -> dict[str, Any] | None:
        rows = await self._request("GET", self.results_base, params={
            "select": "*",
            "id": f"eq.{result_id}",
            "limit": "1",
        })
        return rows[0] if rows else None

    async def insert_game_result(self, row: dict[str, Any]) -> dict[str, Any]:
        rows = await self._request("POST", self.results_base, body=row, prefer="return=representation")
        if not rows:
            raise RuntimeError("Supabase did not return the stored result.")
        return rows[0]

    async def prune_game_results(self, subject_code: str, keep: int) -> int:
        """Delete everything older than the `keep` most recent for this subject."""
        rows = await self._request("GET", self.results_base, params={
            "select": "id,played_at",
            "subject_code": f"eq.{subject_code}",
            "order": "played_at.desc",
            "offset": str(keep),
            "limit": "200",
        })
        removed = 0
        for row in rows or []:
            try:
                await self._request("DELETE", self.results_base, params={"id": f"eq.{row['id']}"})
                removed += 1
            except Exception:
                break
        return removed

    async def get_draft(self, subject_code: str, lecturer_id: str) -> dict[str, Any] | None:
        rows = await self._request("GET", self.drafts_base, params={
            "select": "id,lecturer_id,subject_code,title,chapter,description,questions,question_count,default_time_limit,editing_test_id,updated_at",
            "lecturer_id": f"eq.{lecturer_id}",
            "subject_code": f"eq.{subject_code}",
            "limit": "1",
        })
        return rows[0] if rows else None

    async def save_draft(self, subject_code: str, lecturer: dict[str, Any], payload: DraftPayload) -> dict[str, Any]:
        existing = await self.get_draft(subject_code, lecturer["id"])
        body = {
            "lecturer_id": lecturer["id"],
            "subject_code": subject_code,
            "title": payload.title,
            "chapter": payload.chapter or None,
            "description": payload.description or None,
            "question_count": len(payload.questions),
            "default_time_limit": payload.default_time_limit,
            "questions": [q.model_dump() for q in payload.questions],
            "editing_test_id": payload.editingTestId,
            "owner_name": lecturer.get("name") or lecturer.get("email") or "Lecturer",
            "updated_at": datetime.now(UTC).isoformat(),
        }
        if existing:
            rows = await self._request("PATCH", self.drafts_base, params={
                "id": f"eq.{existing['id']}",
                "lecturer_id": f"eq.{lecturer['id']}",
            }, body=body, prefer="return=representation")
        else:
            rows = await self._request("POST", self.drafts_base, body=body, prefer="return=representation")
        if not rows:
            raise RuntimeError("Supabase did not return the saved draft.")
        return rows[0]

    async def clear_draft(self, subject_code: str, lecturer_id: str) -> None:
        await self._request("DELETE", self.drafts_base, params={
            "lecturer_id": f"eq.{lecturer_id}",
            "subject_code": f"eq.{subject_code}",
        })


class HybridTestRepository:
    def __init__(self, subjects: dict[str, Any]):
        self.subjects = subjects
        self.builtin_tests: dict[str, dict[str, dict[str, Any]]] = {}
        self.local_custom_tests: dict[str, dict[str, dict[str, Any]]] = {}
        self.local_drafts: dict[tuple[str, str], dict[str, Any]] = {}
        self.local_results: dict[str, list[dict[str, Any]]] = {}
        self.results_error: str | None = None
        self.local_lecturers: dict[str, dict[str, Any]] = {}
        self.local_subjects: dict[str, dict[str, Any]] = {}
        self.local_store_path = LOCAL_STORE_PATH
        self.local_store_enabled = True
        self.local_store_error: str | None = None
        self.require_supabase = REQUIRE_SUPABASE
        self.supabase_configured = False
        self.supabase_error: str | None = None
        # Draft failures are tracked separately: they must never disable the
        # Supabase connection used for tests, but they must still be visible.
        self.draft_error: str | None = None
        self._seed_builtin_tests()
        self._load_local_store()

        url = os.environ.get("SUPABASE_URL", "").strip()
        key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "").strip()
        self.supabase_configured = bool(url and key)
        self.remote = SupabaseStore(url, key) if self.supabase_configured else None
        self._set_storage_mode()

    def _set_storage_mode(self) -> None:
        if self.remote is not None:
            self.storage_mode = "supabase"
        elif self.local_store_enabled:
            self.storage_mode = "local-file"
        else:
            self.storage_mode = "in-memory"

    def supabase_unavailable(self) -> bool:
        return self.supabase_configured and self.remote is None and bool(self.supabase_error)

    def _ensure_supabase_for_write(self) -> None:
        if self.require_supabase and self.supabase_configured and self.remote is None:
            raise SupabaseUnavailable("Supabase is unavailable. Writes are disabled while REQUIRE_SUPABASE is enabled.")

    def _draft_key(self, lecturer_id: str, subject_code: str) -> str:
        return f"{lecturer_id}::{subject_code}"

    def _parse_draft_key(self, key: str) -> tuple[str, str] | None:
        if not isinstance(key, str) or "::" not in key:
            return None
        parts = key.split("::", 1)
        if len(parts) != 2 or not parts[0] or not parts[1]:
            return None
        return parts[0], parts[1]

    def _register_subject(self, code: str, name: str) -> None:
        code = (code or "").strip().upper()
        name = (name or "").strip()
        if not code or not name or code in BUILTIN_SUBJECT_CODES:
            return
        entry = self.subjects.get(code)
        if entry:
            entry["name"] = name
            entry.setdefault("questions", [])
        else:
            self.subjects[code] = {"code": code, "name": name, "questions": []}

    def _load_local_store(self) -> None:
        if not self.local_store_path.exists():
            return
        try:
            with self.local_store_path.open("r", encoding="utf-8") as handle:
                data = json.load(handle)
        except Exception as exc:
            self.local_store_error = f"Failed to load local store: {exc}"
            return
        if not isinstance(data, dict):
            return
        subjects = data.get("local_subjects", {})
        if isinstance(subjects, dict):
            for raw_code, row in subjects.items():
                if not isinstance(row, dict):
                    continue
                code = str(row.get("code") or raw_code or "").strip().upper()
                name = str(row.get("name") or "").strip()
                if not code or not name or code in BUILTIN_SUBJECT_CODES:
                    continue
                cleaned = {
                    "code": code,
                    "name": name,
                    "created_by": row.get("created_by"),
                    "created_at": row.get("created_at"),
                }
                self.local_subjects[code] = cleaned
                self._register_subject(code, name)
        tests = data.get("local_custom_tests", {})
        if isinstance(tests, dict):
            for code, items in tests.items():
                if code not in self.subjects or not isinstance(items, dict):
                    continue
                cleaned: dict[str, dict[str, Any]] = {}
                for test_id, row in items.items():
                    if not isinstance(row, dict):
                        continue
                    row.setdefault("subject_code", code)
                    row["source"] = row.get("source") or "local-file"
                    cleaned[test_id] = row
                if cleaned:
                    self.local_custom_tests[code] = cleaned
        drafts = data.get("local_drafts", {})
        if isinstance(drafts, dict):
            for key, row in drafts.items():
                parsed = self._parse_draft_key(key)
                if not parsed or not isinstance(row, dict):
                    continue
                self.local_drafts[parsed] = row
        lecturers = data.get("local_lecturers", {})
        if isinstance(lecturers, dict):
            self.local_lecturers = lecturers
        results = data.get("local_results", {})
        if isinstance(results, dict):
            for code, rows in results.items():
                if isinstance(rows, list):
                    self.local_results[code] = [r for r in rows if isinstance(r, dict)][:RESULTS_RETENTION]

    def _persist_local_store(self) -> None:
        if not self.local_store_enabled:
            return
        payload = {
            "version": LOCAL_STORE_VERSION,
            "local_subjects": self.local_subjects,
            "local_custom_tests": self.local_custom_tests,
            "local_drafts": {
                self._draft_key(lecturer_id, subject_code): row
                for (lecturer_id, subject_code), row in self.local_drafts.items()
            },
            "local_lecturers": self.local_lecturers,
            "local_results": self.local_results,
        }
        tmp_path = self.local_store_path.with_suffix(".tmp")
        try:
            with tmp_path.open("w", encoding="utf-8") as handle:
                json.dump(payload, handle, ensure_ascii=True, indent=2, sort_keys=True)
            tmp_path.replace(self.local_store_path)
        except Exception as exc:
            self.local_store_error = f"Failed to save local store: {exc}"
            self.local_store_enabled = False
            self._set_storage_mode()

    def _cache_lecturer_row(self, row: dict[str, Any] | None) -> None:
        if not row or "password_hash" not in row:
            return
        email = (row.get("email") or "").strip().lower()
        if not email:
            return
        existing = self.local_lecturers.get(email, {})
        merged = dict(existing)
        merged.update(row)
        self.local_lecturers[email] = merged
        self._persist_local_store()

    def _handle_supabase_error(self, exc: RuntimeError) -> bool:
        message = str(exc)
        message_lower = message.lower()
        if (
            "pgrst205" in message_lower
            or "schema cache" in message_lower
            or ("could not find the table" in message_lower)
            or ("relation" in message_lower and "does not exist" in message_lower)
        ):
            self.supabase_error = message
            self.remote = None
            self.local_store_enabled = True
            self._set_storage_mode()
            return True
        return False

    async def _call_remote(self, awaitable, fallback):
        if self.remote is None or awaitable is None:
            return fallback()
        try:
            return await awaitable
        except RuntimeError as exc:
            if self._handle_supabase_error(exc):
                return fallback()
            raise
        except httpx.RequestError as exc:
            self.supabase_error = str(exc)
            self.remote = None
            self.local_store_enabled = True
            self._set_storage_mode()
            return fallback()

    def _seed_builtin_tests(self) -> None:
        for code in self.subjects:
            self.builtin_tests[code] = {}
            self.local_custom_tests[code] = {}

    def get_storage_status(self) -> dict[str, Any]:
        if self.remote is None:
            if self.local_store_enabled:
                if self.supabase_configured and self.supabase_error:
                    note = "Supabase configured but schema is missing. Using local file storage."
                else:
                    note = "Local file storage is active on this server. Data resets on redeploy unless a persistent disk is used."
            else:
                if self.supabase_configured and self.supabase_error:
                    note = "Supabase configured but schema is missing. Running in-memory until the schema is applied."
                else:
                    note = "In-memory storage resets on redeploy/restart."
        else:
            note = "Supabase storage is active."
        return {
            "mode": self.storage_mode,
            "supabaseConfigured": self.supabase_configured,
            "note": note,
            "supabaseError": self.supabase_error,
            "asleep": self.supabase_looks_asleep(),
            "results": self.results_storage_usage(),
        }

    def supabase_looks_asleep(self) -> bool:
        """Detect a paused Supabase free project.

        Free projects are paused after 7 days of inactivity — which lands in the
        middle of a semester break — and then return connection failures rather
        than anything explanatory. Better to say "the database is asleep, resume
        it in your dashboard" than to show a generic error.
        """
        if not self.supabase_configured:
            return False
        blob = f"{self.supabase_error or ''} {self.results_error or ''} {self.draft_error or ''}".lower()
        if not blob.strip():
            return False
        return any(
            marker in blob
            for marker in (
                "getaddrinfo", "name or service not known", "nodename nor servname",
                "connection refused", "connect call failed", "timed out", "timeout",
                "connecterror", "temporary failure in name resolution",
                "project is paused", "service unavailable", "502", "503",
            )
        )

    def _summary(self, row: dict[str, Any], lecturer_id: str | None = None) -> dict[str, Any]:
        created_by = row.get("created_by")
        source = row.get("source", "supabase")
        can_edit = False
        if source not in {"built-in"}:
            can_edit = bool(lecturer_id and created_by and created_by == lecturer_id)
        return {
            "id": row["id"],
            "subject_code": row["subject_code"],
            "title": row.get("title", "Untitled Test"),
            "chapter": row.get("chapter") or "",
            "description": row.get("description") or "",
            "questionCount": row.get("question_count") or len(row.get("questions") or []),
            "defaultTimeLimit": coerce_time_limit(row.get("default_time_limit")) or TIME_PER_Q,
            "estimatedSeconds": estimate_test_seconds(row),
            "source": source,
            "created_at": row.get("created_at"),
            "updated_at": row.get("updated_at"),
            "ownerName": row.get("owner_name") or "System",
            "createdBy": created_by,
            "canEdit": can_edit,
        }

    async def get_lecturer_by_email(self, email: str) -> dict[str, Any] | None:
        email = email.strip().lower()
        def _local_lookup():
            return self.local_lecturers.get(email)
        if self.remote is None:
            return _local_lookup()
        try:
            row = await self.remote.get_lecturer_by_email(email)
        except RuntimeError as exc:
            if self._handle_supabase_error(exc):
                return _local_lookup()
            raise
        except httpx.RequestError as exc:
            self.supabase_error = str(exc)
            self.remote = None
            self.local_store_enabled = True
            self._set_storage_mode()
            return _local_lookup()
        if row:
            self._cache_lecturer_row(row)
            return row
        return _local_lookup()

    async def get_lecturer_by_id(self, lecturer_id: str) -> dict[str, Any] | None:
        def _local_lookup():
            for row in self.local_lecturers.values():
                if row["id"] == lecturer_id:
                    return {k: v for k, v in row.items() if k != "password_hash"}
            return None
        if self.remote is None:
            return _local_lookup()
        try:
            row = await self.remote.get_lecturer_by_id(lecturer_id)
        except RuntimeError as exc:
            if self._handle_supabase_error(exc):
                return _local_lookup()
            raise
        except httpx.RequestError as exc:
            self.supabase_error = str(exc)
            self.remote = None
            self.local_store_enabled = True
            self._set_storage_mode()
            return _local_lookup()
        return row or _local_lookup()

    async def create_lecturer(self, payload: LecturerSignupPayload) -> dict[str, Any]:
        self._ensure_supabase_for_write()
        if await self.get_lecturer_by_email(payload.email):
            raise ValueError("An account with that email already exists.")
        password_hash = await hash_password_async(payload.password)
        def _local_create():
            row = {
                "id": str(uuid.uuid4()),
                "name": payload.name,
                "email": payload.email,
                "password_hash": password_hash,
                "created_at": datetime.now(UTC).isoformat(),
                "updated_at": datetime.now(UTC).isoformat(),
            }
            self.local_lecturers[payload.email] = row
            self._persist_local_store()
            return {k: v for k, v in row.items() if k != "password_hash"}
        result = await self._call_remote(
            self.remote.create_lecturer(payload.name, payload.email, password_hash) if self.remote else None,
            _local_create
        )
        self._cache_lecturer_row(result)
        return result

    async def list_subjects(self) -> list[dict[str, Any]]:
        remote_rows = await self._call_remote(
            self.remote.list_subjects() if self.remote else None,
            lambda: list(self.local_subjects.values())
        )
        combined: dict[str, dict[str, Any]] = {}
        for row in list(self.local_subjects.values()):
            code = (row.get("code") or "").strip().upper()
            name = (row.get("name") or "").strip()
            if not code or not name or code in BUILTIN_SUBJECT_CODES:
                continue
            combined[code] = {
                "code": code,
                "name": name,
                "created_by": row.get("created_by"),
                "created_at": row.get("created_at"),
            }
            self._register_subject(code, name)
        for row in remote_rows or []:
            code = (row.get("code") or "").strip().upper()
            name = (row.get("name") or "").strip()
            if not code or not name or code in BUILTIN_SUBJECT_CODES:
                continue
            combined[code] = {
                "code": code,
                "name": name,
                "created_by": row.get("created_by"),
                "created_at": row.get("created_at"),
            }
            self._register_subject(code, name)
        if combined:
            self.local_subjects.update(combined)
            self._persist_local_store()
        return list(combined.values())

    async def create_subject(self, code: str, name: str, lecturer: dict[str, Any]) -> dict[str, Any]:
        self._ensure_supabase_for_write()
        code = (code or "").strip().upper()
        name = (name or "").strip()
        if code in BUILTIN_SUBJECT_CODES or code in self.subjects:
            raise ValueError("Subject code already exists.")
        if self.remote is not None:
            existing = await self._call_remote(self.remote.get_subject(code), lambda: None)
            if existing:
                raise ValueError("Subject code already exists.")
        def _local_create():
            row = {
                "code": code,
                "name": name,
                "created_by": lecturer.get("id"),
                "created_at": datetime.now(UTC).isoformat(),
            }
            self.local_subjects[code] = row
            self._register_subject(code, name)
            self._persist_local_store()
            return row
        async def _remote_create():
            try:
                return await self.remote.create_subject(code, name, lecturer["id"])
            except RuntimeError as exc:
                if "duplicate" in str(exc).lower():
                    raise ValueError("Subject code already exists.") from exc
                raise
        row = await self._call_remote(_remote_create() if self.remote else None, _local_create)
        if row:
            self.local_subjects[code] = {
                "code": code,
                "name": row.get("name") or name,
                "created_by": row.get("created_by") or lecturer.get("id"),
                "created_at": row.get("created_at"),
            }
            self._register_subject(code, row.get("name") or name)
            self._persist_local_store()
        return row

    async def delete_subject(self, code: str, lecturer: dict[str, Any]) -> dict[str, Any]:
        self._ensure_supabase_for_write()
        code = (code or "").strip().upper()
        if code in BUILTIN_SUBJECT_CODES:
            raise PermissionError("Built-in subjects cannot be deleted.")

        def _local_delete():
            row = self.local_subjects.get(code)
            if not row:
                raise KeyError("Subject not found")
            if row.get("created_by") != lecturer.get("id"):
                raise PermissionError("Only the lecturer who created this subject can delete it.")
            if self.local_custom_tests.get(code):
                raise ValueError("Cannot delete a subject with saved tests.")
            self.local_subjects.pop(code, None)
            if code in self.subjects and code not in BUILTIN_SUBJECT_CODES:
                self.subjects.pop(code, None)
            self.local_custom_tests.pop(code, None)
            self._persist_local_store()
            return row

        async def _remote_delete():
            row = await self.remote.get_subject(code)
            if not row:
                raise KeyError("Subject not found")
            if row.get("created_by") != lecturer.get("id"):
                raise PermissionError("Only the lecturer who created this subject can delete it.")
            if await self.remote.subject_has_tests(code) or self.local_custom_tests.get(code):
                raise ValueError("Cannot delete a subject with saved tests.")
            await self.remote.delete_subject(code, lecturer.get("id"))
            return row

        row = await self._call_remote(_remote_delete() if self.remote else None, _local_delete)
        self.local_subjects.pop(code, None)
        if code in self.subjects and code not in BUILTIN_SUBJECT_CODES:
            self.subjects.pop(code, None)
        self.local_custom_tests.pop(code, None)
        self._persist_local_store()
        return row

    async def list_tests_by_creator(self, lecturer_id: str) -> list[dict[str, Any]]:
        remote_rows = await self._call_remote(
            self.remote.list_tests_by_creator(lecturer_id) if self.remote else None,
            lambda: []
        )
        local_rows: list[dict[str, Any]] = []
        for items in self.local_custom_tests.values():
            for row in items.values():
                if row.get("created_by") == lecturer_id:
                    local_rows.append(row)
        return list(remote_rows or []) + local_rows

    async def get_test_counts(self) -> dict[str, dict[str, int]]:
        """Test and question counts per subject, in a single Supabase round trip.

        Replaces one list_tests() call per subject on GET /api/subjects.
        """
        counts: dict[str, dict[str, int]] = {}

        def _add(subject_code: str, question_count: int) -> None:
            entry = counts.setdefault(subject_code, {"tests": 0, "questions": 0})
            entry["tests"] += 1
            entry["questions"] += question_count

        remote_rows = await self._call_remote(
            self.remote.list_all_test_counts() if self.remote else None,
            lambda: []
        )
        for row in remote_rows or []:
            code = (row.get("subject_code") or "").strip().upper()
            if code:
                _add(code, int(row.get("question_count") or 0))
        for code, items in self.local_custom_tests.items():
            for row in items.values():
                _add(code, int(row.get("question_count") or len(row.get("questions") or [])))
        return counts

    async def list_tests(self, subject_code: str, lecturer_id: str | None = None) -> list[dict[str, Any]]:
        if subject_code not in self.subjects:
            raise KeyError(subject_code)

        tests: list[dict[str, Any]] = []
        remote_rows = await self._call_remote(
            self.remote.list_tests(subject_code, lecturer_id) if self.remote else None,
            lambda: []
        )
        local_rows = list(self.local_custom_tests.get(subject_code, {}).values())

        tests.extend(self._summary(row, lecturer_id) for row in remote_rows)
        tests.extend(self._summary(row, lecturer_id) for row in local_rows)
        return tests

    async def get_test(self, subject_code: str, test_id: str, lecturer_id: str | None = None) -> dict[str, Any] | None:
        if test_id in self.builtin_tests.get(subject_code, {}):
            row = self.builtin_tests[subject_code][test_id]
            row["can_edit"] = False
            return row
        if test_id in self.local_custom_tests.get(subject_code, {}):
            row = self.local_custom_tests[subject_code][test_id]
            row["can_edit"] = bool(lecturer_id and row.get("created_by") == lecturer_id)
            return row
        return await self._call_remote(
            self.remote.get_test(subject_code, test_id, lecturer_id) if self.remote else None,
            lambda: None
        )

    async def create_test(self, subject_code: str, payload: TestPayload, lecturer: dict[str, Any]) -> dict[str, Any]:
        self._ensure_supabase_for_write()
        if subject_code not in self.subjects:
            raise KeyError(subject_code)
        def _local_create():
            test_id = f"local:{uuid.uuid4()}"
            row = {
                "id": test_id,
                "subject_code": subject_code,
                "title": payload.title,
                "chapter": payload.chapter,
                "description": payload.description,
                "question_count": len(payload.questions),
                "default_time_limit": payload.default_time_limit,
                "questions": [q.model_dump() for q in payload.questions],
                "created_at": datetime.now(UTC).isoformat(),
                "updated_at": datetime.now(UTC).isoformat(),
                "source": "local-file" if self.local_store_enabled else "in-memory",
                "created_by": lecturer["id"],
                "owner_name": lecturer.get("name") or lecturer.get("email") or "Lecturer",
            }
            self.local_custom_tests.setdefault(subject_code, {})[test_id] = row
            self._persist_local_store()
            return row
        async def _remote_create():
            row = await self.remote.create_test(subject_code, payload, lecturer)
            row.setdefault("question_count", len(payload.questions))
            row.setdefault("questions", [q.model_dump() for q in payload.questions])
            row.setdefault("owner_name", lecturer.get("name") or lecturer.get("email") or "Lecturer")
            row.setdefault("created_by", lecturer["id"])
            return row
        return await self._call_remote(_remote_create() if self.remote else None, _local_create)

    async def update_test(self, subject_code: str, test_id: str, payload: TestPayload, lecturer: dict[str, Any]) -> dict[str, Any]:
        self._ensure_supabase_for_write()
        if test_id in self.builtin_tests.get(subject_code, {}):
            raise PermissionError("Built-in tests cannot be edited.")
        def _local_update():
            row = self.local_custom_tests.get(subject_code, {}).get(test_id)
            if not row:
                raise KeyError("Test not found")
            if row.get("created_by") and row.get("created_by") != lecturer["id"]:
                raise PermissionError("Only the lecturer who created this test can edit it.")
            row.update({
                "title": payload.title,
                "chapter": payload.chapter,
                "description": payload.description,
                "question_count": len(payload.questions),
                "default_time_limit": payload.default_time_limit,
                "questions": [q.model_dump() for q in payload.questions],
                "updated_at": datetime.now(UTC).isoformat(),
            })
            self._persist_local_store()
            return row
        async def _remote_update():
            row = await self.remote.update_test(subject_code, test_id, payload, lecturer)
            row.setdefault("question_count", len(payload.questions))
            row.setdefault("questions", [q.model_dump() for q in payload.questions])
            row.setdefault("owner_name", lecturer.get("name") or lecturer.get("email") or "Lecturer")
            row.setdefault("created_by", lecturer["id"])
            return row
        return await self._call_remote(_remote_update() if self.remote else None, _local_update)

    async def delete_test(self, subject_code: str, test_id: str, lecturer: dict[str, Any]) -> dict[str, Any]:
        self._ensure_supabase_for_write()
        if test_id in self.builtin_tests.get(subject_code, {}):
            raise PermissionError("Built-in tests cannot be deleted.")

        def _local_delete():
            row = self.local_custom_tests.get(subject_code, {}).get(test_id)
            if not row:
                raise KeyError("Test not found")
            if row.get("created_by") != lecturer["id"]:
                raise PermissionError("Only the lecturer who created this test can delete it.")
            self.local_custom_tests.get(subject_code, {}).pop(test_id, None)
            self._persist_local_store()
            return row

        async def _remote_delete():
            existing = await self.remote.get_test(subject_code, test_id, lecturer.get("id"))
            if not existing:
                raise KeyError("Test not found")
            if existing.get("created_by") != lecturer["id"]:
                raise PermissionError("Only the lecturer who created this test can delete it.")
            await self.remote.delete_test(subject_code, test_id)
            return existing

        return await self._call_remote(_remote_delete() if self.remote else None, _local_delete)

    # ── Stored game results ──────────────────────────────────────────────────
    # Deliberately lightweight. The automatic spreadsheet download at the end of
    # a game stays the authoritative record; this exists so a Render redeploy
    # does not destroy results that were never downloaded. Question text is NOT
    # duplicated — the test id is referenced instead.

    def _result_row(self, stats: dict[str, Any], lecturer_id: str | None) -> dict[str, Any]:
        players = []
        for vid, player in (stats.get("players") or {}).items():
            players.append({
                "id": vid,
                "name": player.get("name", ""),
                "student_number": player.get("student_number", ""),
                "score": player.get("score", 0),
                # Only what is needed to rebuild the spreadsheet against the test.
                "answers": [
                    {
                        "q": a.get("q"),
                        "choice": a.get("choice", -1),
                        "correct": bool(a.get("correct")),
                        "points": a.get("points", 0),
                        "time": round(float(a.get("time") or 0), 2),
                    }
                    for a in (player.get("answers") or [])
                ],
            })
        return {
            "subject_code": stats.get("subject_code"),
            "test_id": stats.get("test_id"),
            "test_title": stats.get("test_title") or "",
            "session_name": stats.get("session_name") or "",
            "played_at": stats.get("timestamp") or datetime.now(UTC).isoformat(),
            "player_count": len(players),
            "question_count": len(stats.get("questions") or []),
            "players": players,
            "created_by": lecturer_id,
        }

    async def store_game_result(self, stats: dict[str, Any], lecturer_id: str | None = None) -> dict[str, Any] | None:
        if not PERSIST_RESULTS or not stats or not stats.get("players"):
            return None
        row = self._result_row(stats, lecturer_id)
        subject_code = row["subject_code"]

        def _local_store():
            row_local = dict(row, id=f"result:{uuid.uuid4()}")
            bucket = self.local_results.setdefault(subject_code, [])
            bucket.append(row_local)
            bucket.sort(key=lambda item: str(item.get("played_at") or ""), reverse=True)
            del bucket[RESULTS_RETENTION:]          # prune after each insert
            self._persist_local_store()
            return row_local

        if self.remote is not None:
            try:
                stored = await self.remote.insert_game_result(row)
                try:
                    await self.remote.prune_game_results(subject_code, RESULTS_RETENTION)
                except Exception as exc:
                    print(f"[results] prune failed for {subject_code}: {exc}")
                return stored
            except Exception as exc:
                # Never let this break the end of a game — the lecturer's
                # automatic download is the record that matters.
                self.results_error = str(exc)
                print(f"[results] Supabase store failed for {subject_code}: {exc}")
        try:
            return _local_store()
        except Exception as exc:
            self.results_error = str(exc)
            return None

    async def list_game_results(self, subject_code: str) -> list[dict[str, Any]]:
        remote_rows = []
        if self.remote is not None:
            try:
                remote_rows = await self.remote.list_game_results(subject_code, RESULTS_RETENTION)
            except Exception as exc:
                self.results_error = str(exc)
        local_rows = [
            {k: v for k, v in row.items() if k != "players"}
            for row in self.local_results.get(subject_code, [])
        ]
        combined = list(remote_rows or []) + local_rows
        combined.sort(key=lambda item: str(item.get("played_at") or ""), reverse=True)
        return combined[:RESULTS_RETENTION]

    async def get_game_result(self, subject_code: str, result_id: str) -> dict[str, Any] | None:
        for row in self.local_results.get(subject_code, []):
            if row.get("id") == result_id:
                return row
        if self.remote is not None:
            try:
                return await self.remote.get_game_result(result_id)
            except Exception as exc:
                self.results_error = str(exc)
        return None

    def results_storage_usage(self, subject_code: str | None = None) -> dict[str, Any]:
        """Roughly how much space stored results occupy, so the cost is visible."""
        buckets = (
            [self.local_results.get(subject_code, [])] if subject_code
            else list(self.local_results.values())
        )
        sessions = sum(len(bucket) for bucket in buckets)
        approx_bytes = sum(
            len(json.dumps(row, ensure_ascii=False).encode("utf-8"))
            for bucket in buckets for row in bucket
        )
        return {
            "enabled": PERSIST_RESULTS,
            "retention": RESULTS_RETENTION,
            "sessions": sessions,
            "approxBytes": approx_bytes,
            "approxKb": round(approx_bytes / 1024, 1),
        }

    def _local_draft_backend(self) -> str:
        return "local-file" if self.local_store_enabled else "memory"

    async def get_draft(self, subject_code: str, lecturer: dict[str, Any]) -> tuple[dict[str, Any] | None, str, str | None]:
        """Return (draft, backend_it_came_from, error_text).

        Do NOT use _call_remote here — a failure on the quiz_test_drafts table
        must never disable self.remote, which would break all subsequent test
        reads and writes.
        """
        error: str | None = None
        if self.remote is not None:
            try:
                row = await self.remote.get_draft(subject_code, lecturer["id"])
                if row is not None:
                    return row, "supabase", None
                # Supabase is reachable and simply has no draft. Still fall through
                # to the local copy: a draft written while Supabase was down lives
                # there and would otherwise be invisible.
            except Exception as exc:
                error = str(exc)
                self.draft_error = error
        local = self.local_drafts.get((lecturer["id"], subject_code))
        if local is not None:
            return local, self._local_draft_backend(), error
        return None, "supabase" if (self.remote is not None and not error) else self._local_draft_backend(), error

    async def save_draft(self, subject_code: str, lecturer: dict[str, Any], payload: DraftPayload) -> tuple[dict[str, Any], str, str | None]:
        """Save a draft and report which backend actually stored it.

        Returns (row, backend, supabase_error). Raises if nothing could store it
        — a draft the lecturer believes is safe but which was never written is
        worse than an error message.
        """
        self._ensure_supabase_for_write()

        def _local_save():
            row = {
                "id": self.local_drafts.get((lecturer["id"], subject_code), {}).get("id", f"draft:{uuid.uuid4()}"),
                "lecturer_id": lecturer["id"],
                "subject_code": subject_code,
                "title": payload.title,
                "chapter": payload.chapter,
                "description": payload.description,
                "question_count": len(payload.questions),
                "default_time_limit": payload.default_time_limit,
                "questions": [q.model_dump() for q in payload.questions],
                "editing_test_id": payload.editingTestId,
                "updated_at": datetime.now(UTC).isoformat(),
                "owner_name": lecturer.get("name") or lecturer.get("email") or "Lecturer",
            }
            self.local_drafts[(lecturer["id"], subject_code)] = row
            self._persist_local_store()
            return row

        remote_error: str | None = None
        if self.remote is not None:
            try:
                row = await self.remote.save_draft(subject_code, lecturer, payload)
                self.draft_error = None
                # Mirror into the local store so a later Supabase outage still has
                # something to hand back.
                try:
                    _local_save()
                except Exception:
                    pass
                return row, "supabase", None
            except Exception as exc:
                # Log it. The old code swallowed this entirely, so a draft could
                # land in an ephemeral file while the UI said "Draft saved".
                remote_error = str(exc)
                self.draft_error = remote_error
                print(f"[drafts] Supabase draft save failed for {subject_code}: {remote_error}")

        row = _local_save()
        # _persist_local_store() flips local_store_enabled off if the file could
        # not be written, so read the backend back after saving, not before.
        return row, self._local_draft_backend(), remote_error

    async def clear_draft(self, subject_code: str, lecturer: dict[str, Any]) -> None:
        # Always clear the local draft copy first.
        self.local_drafts.pop((lecturer["id"], subject_code), None)
        self._persist_local_store()
        # Do NOT use _call_remote here — a failure on the quiz_test_drafts table must
        # never disable self.remote (which would break all subsequent test reads/writes).
        if self.remote is not None:
            try:
                await self.remote.clear_draft(subject_code, lecturer["id"])
            except Exception:
                pass  # Non-critical; the test itself was already saved successfully


repo = HybridTestRepository(SUBJECTS)


# ──────────────────────────────────────────────────────────────────────────────
# Per-room live game state
# ──────────────────────────────────────────────────────────────────────────────
def coerce_time_limit(value: Any) -> int | None:
    """A usable limit within bounds, or None to fall through to the next level."""
    if value is None or isinstance(value, bool):
        return None
    try:
        seconds = int(value)
    except (TypeError, ValueError):
        return None
    if MIN_TIME_LIMIT <= seconds <= MAX_TIME_LIMIT:
        return seconds
    return None


def estimate_test_seconds(row: dict[str, Any]) -> int:
    """Rough wall-clock length: each question's limit plus the ready and reveal pauses."""
    default_limit = coerce_time_limit(row.get("default_time_limit")) or TIME_PER_Q
    questions = row.get("questions") or []
    if not questions:
        count = int(row.get("question_count") or 0)
        return int(count * (default_limit + GET_READY_SECONDS + REVEAL_SECONDS))
    total = 0.0
    for question in questions:
        limit = default_limit
        if isinstance(question, dict):
            limit = coerce_time_limit(question.get("time_limit")) or default_limit
        total += limit + GET_READY_SECONDS + REVEAL_SECONDS
    return int(total)


MAX_PLAYERS_PER_ROOM = int(os.environ.get("MAX_PLAYERS_PER_ROOM", "300"))
MAX_WS_MESSAGE_BYTES = int(os.environ.get("MAX_WS_MESSAGE_BYTES", str(64 * 1024)))


class GameRoom:
    def __init__(self, subject_code: str):
        self.subject_code = subject_code
        subject = SUBJECTS.get(subject_code)
        if subject is None:
            # A room for a subject that is not in the catalogue used to raise
            # KeyError and take down whatever was constructing it.
            self.subject_name = subject_code
        else:
            self.subject_name = subject.get("name") or subject_code
        # Serialises question advance / reveal so a double click on "Next
        # Question", or a click racing the auto-reveal timer, cannot increment
        # current_q twice and skip a question.
        self.advance_lock = asyncio.Lock()
        self.last_game_stats = None
        self.active_test_id = None
        self.active_test_title = ""
        self.active_test_chapter = ""
        self.questions: list[dict[str, Any]] = []
        self.total_q = 0
        self.session_name = ""
        self.current_token = ""
        self.game_code = ""
        self.game_code_enabled = False
        self.default_time_limit = TIME_PER_Q
        self.reset_runtime_state(clear_players=True)

    def set_active_test(self, test_data: dict[str, Any] | None) -> None:
        self.active_test_id = test_data.get("id") if test_data else None
        self.active_test_title = test_data.get("title", "") if test_data else ""
        self.active_test_chapter = test_data.get("chapter", "") if test_data else ""
        self.questions = list(test_data.get("questions", [])) if test_data else []
        self.total_q = len(self.questions)
        self.default_time_limit = coerce_time_limit(
            test_data.get("default_time_limit") if test_data else None
        ) or TIME_PER_Q

    def time_limit_for(self, index: int | None = None) -> int:
        """Resolve the limit: question level → test level → TIME_PER_Q.

        Tests saved before this feature have neither, so they keep running at
        exactly 30 seconds with no migration.
        """
        if index is None:
            index = self.current_q
        if 0 <= index < len(self.questions):
            question = self.questions[index]
            if isinstance(question, dict):
                per_question = coerce_time_limit(question.get("time_limit"))
                if per_question:
                    return per_question
        return coerce_time_limit(getattr(self, "default_time_limit", None)) or TIME_PER_Q

    def reset_runtime_state(self, *, clear_players: bool) -> None:
        self.phase = "lobby"
        self.current_q = 0
        self.question_start_time = 0
        self.question_elapsed = 0.0
        self.game_code = ""
        self.game_code_enabled = False
        if clear_players:
            self.players = {}
            # Everyone who has taken part in this session, keyed by student
            # identity. Kept even if their live record goes away, so a student
            # who drops out mid-quiz is never mistaken for a new joiner.
            self.session_roster = {}
            self.current_token = ""
        self.host_ws = None
        self.host_visitor = None
        self.host_lecturer_id: str | None = None
        self.answers_this_round = {}
        self.question_timer_task = None
        self.question_time_limit = TIME_PER_Q
        self.start_task: asyncio.Task | None = None
        self.paused = False
        self._pending_room_update: asyncio.Task | None = None

    def archive_stats(self) -> None:
        if not self.players:
            return
        self.last_game_stats = {
            "subject_code": self.subject_code,
            "subject_name": self.subject_name,
            "test_id": self.active_test_id,
            "test_title": self.active_test_title,
            "test_chapter": self.active_test_chapter,
            "session_name": self.session_name,
            "timestamp": datetime.now().astimezone().isoformat(),
            "questions": self.questions,
            "players": {}
        }
        for vid, p in self.players.items():
            self.last_game_stats["players"][vid] = {
                "name": p["name"],
                "student_number": p.get("student_number", ""),
                "score": p["score"],
                "answers": p["answers"]
            }


rooms: dict[str, GameRoom] = {code: GameRoom(code) for code in SUBJECTS}
session_tokens: dict[str, dict[str, Any]] = {}
SESSION_TOKEN_TTL = 60 * 60 * 24 * 7  # 7 days — links should not expire during a quiz session
SESSION_TOKEN_LENGTH = 6
SESSION_TOKEN_ALPHABET = string.ascii_uppercase + string.digits


def consume_session_token(token: str) -> None:
    normalized = (token or "").strip().upper()
    if not normalized:
        return
    session_tokens.pop(normalized, None)


def generate_session_token(subject_code: str) -> str:
    normalized = (subject_code or "").strip().upper()
    to_remove = [token for token, entry in list(session_tokens.items()) if entry.get("subject_code") == normalized]
    for token in to_remove:
        consume_session_token(token)
    token = ""
    while not token or token in session_tokens:
        token = "".join(secrets.choice(SESSION_TOKEN_ALPHABET) for _ in range(SESSION_TOKEN_LENGTH))
    session_tokens[token] = {
        "subject_code": normalized,
        "expires_at": time.time() + SESSION_TOKEN_TTL
    }
    return token


def lookup_session_token(token: str) -> str | None:
    normalized = (token or "").strip().upper()
    if not normalized:
        return None
    entry = session_tokens.get(normalized)
    if not entry:
        return None
    if time.time() > float(entry.get("expires_at") or 0):
        consume_session_token(normalized)
        return None
    subject_code = (entry.get("subject_code") or "").strip().upper()
    return subject_code or None


def get_room_active_token(room: GameRoom | None) -> str:
    if room is None:
        return ""
    token = (room.current_token or "").strip().upper()
    if not token:
        return ""
    if lookup_session_token(token) == room.subject_code:
        return token
    return ""


def generate_game_code() -> str:
    """Generate a random 4-digit numeric code."""
    return f"{secrets.randbelow(9000) + 1000}"


async def _cleanup_expired_tokens() -> None:
    while True:
        await asyncio.sleep(60)
        now = time.time()
        expired = [token for token, entry in list(session_tokens.items()) if now > float(entry.get("expires_at") or 0)]
        for token in expired:
            consume_session_token(token)


def register_subject_in_catalog(code: str, name: str) -> None:
    code = (code or "").strip().upper()
    name = (name or "").strip()
    if not code or not name or code in BUILTIN_SUBJECT_CODES:
        return
    entry = SUBJECTS.get(code)
    if entry:
        entry["name"] = name
        entry.setdefault("questions", [])
    else:
        SUBJECTS[code] = {"code": code, "name": name, "questions": []}
    room = rooms.get(code)
    if room:
        room.subject_name = SUBJECTS[code]["name"]
    else:
        rooms[code] = GameRoom(code)


# ──────────────────────────────────────────────────────────────────────────────
# FastAPI
# ──────────────────────────────────────────────────────────────────────────────
@asynccontextmanager
async def lifespan(app: FastAPI):
    try:
        subjects = await repo.list_subjects()
        for row in subjects:
            register_subject_in_catalog(row.get("code"), row.get("name"))
    except Exception as exc:
        print(f"Failed to load subjects from Supabase: {exc}")
    cleanup_task = asyncio.create_task(_cleanup_expired_tokens())
    try:
        yield
    finally:
        cleanup_task.cancel()
        try:
            await cleanup_task
        except asyncio.CancelledError:
            pass
        if repo.remote is not None:
            try:
                await repo.remote.aclose()
            except Exception:
                pass

_check_session_secret_configured()

app = FastAPI(lifespan=lifespan)

# The app is served same-origin, so it needs no cross-origin access at all.
# Defaulting to "*" meant any site could call the API from a visitor's browser.
# "*" is now opt-in; set ALLOWED_ORIGINS explicitly if you ever need it.
ALLOWED_ORIGINS = os.environ.get("ALLOWED_ORIGINS", "").strip()
if ALLOWED_ORIGINS == "*":
    allow_origins = ["*"]
elif ALLOWED_ORIGINS:
    allow_origins = [origin.strip() for origin in ALLOWED_ORIGINS.split(",") if origin.strip()]
else:
    render_host = os.environ.get("RENDER_EXTERNAL_URL", "").strip()
    allow_origins = [render_host] if render_host else []
app.add_middleware(
    CORSMiddleware,
    allow_origins=allow_origins,
    allow_credentials=bool(allow_origins) and allow_origins != ["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

BASE_DIR = Path(__file__).resolve().parent
app.mount("/static", StaticFiles(directory=BASE_DIR), name="static")


@app.get("/")
def index():
    return FileResponse(BASE_DIR / "index.html")


@app.get("/style.css")
def style_css():
    return FileResponse(BASE_DIR / "style.css", media_type="text/css")


@app.get("/app.js")
def app_js():
    return FileResponse(BASE_DIR / "app.js", media_type="application/javascript")


@app.get("/draft_utils.js")
def draft_utils_js():
    return FileResponse(BASE_DIR / "draft_utils.js", media_type="application/javascript")


def public_lecturer_view(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": row.get("id"),
        "name": row.get("name") or "Lecturer",
        "email": row.get("email"),
    }


def set_session_cookie(response: JSONResponse, lecturer_id: str, request: Request) -> None:
    response.set_cookie(
        SESSION_COOKIE_NAME,
        create_session_token(lecturer_id),
        httponly=True,
        samesite="lax",
        secure=request.url.scheme == "https",
        max_age=SESSION_MAX_AGE,
        path="/",
    )


def clear_session_cookie(response: JSONResponse) -> None:
    response.delete_cookie(SESSION_COOKIE_NAME, path="/")


async def current_lecturer_from_request(request: Request) -> dict[str, Any] | None:
    lecturer_id = parse_session_token(request.cookies.get(SESSION_COOKIE_NAME))
    if not lecturer_id:
        return None
    return await repo.get_lecturer_by_id(lecturer_id)


async def require_lecturer(request: Request) -> dict[str, Any]:
    lecturer = await current_lecturer_from_request(request)
    if not lecturer:
        raise HTTPException(status_code=401, detail="Lecturer sign-in required")
    return lecturer


def validation_detail(exc: ValidationError) -> list[dict[str, Any]]:
    """JSON-safe validation errors.

    Pydantic's `errors()` puts the original exception object in `ctx`, which
    FastAPI cannot serialise — the response blows up with a TypeError and the
    lecturer sees a dead request instead of "Each question must have exactly 4
    options." Keep only the fields the front end actually reads.
    """
    detail = []
    for error in exc.errors():
        detail.append({
            "loc": [str(part) for part in error.get("loc", ())],
            "msg": str(error.get("msg", "Invalid value")),
            "type": str(error.get("type", "value_error")),
        })
    return detail


async def current_lecturer_from_websocket(websocket: WebSocket) -> dict[str, Any] | None:
    lecturer_id = parse_session_token(websocket.cookies.get(SESSION_COOKIE_NAME))
    if not lecturer_id:
        return None
    return await repo.get_lecturer_by_id(lecturer_id)


def public_storage_status(status: dict[str, Any]) -> dict[str, Any]:
    """Storage status without the raw Supabase error text.

    The detail names tables, schema state and sometimes the project URL. Only
    signed-in lecturers see it.
    """
    return {
        "mode": status.get("mode"),
        "supabaseConfigured": status.get("supabaseConfigured"),
        "note": status.get("note"),
        "healthy": not status.get("supabaseError"),
        "asleep": status.get("asleep", False),
    }


@app.get("/api/health")
async def health(request: Request):
    status = repo.get_storage_status()
    lecturer = await current_lecturer_from_request(request)
    return {"ok": True, "storage": status if lecturer else public_storage_status(status)}


@app.get("/api/storage-status")
async def storage_status(request: Request):
    status = repo.get_storage_status()
    lecturer = await current_lecturer_from_request(request)
    return status if lecturer else public_storage_status(status)


@app.post("/api/visitor-token")
@limiter.limit("60/minute")
async def issue_visitor_token(request: Request):
    """Issue a signed student identity.

    Students never choose their own id: the server mints it, signs it, and only
    accepts it back inside the signature.
    """
    existing = parse_visitor_token((await _read_json_body(request)).get("token"))
    visitor_id, token = create_visitor_token(existing)
    return {"visitorId": visitor_id, "token": token}


async def _read_json_body(request: Request) -> dict[str, Any]:
    try:
        body = await request.json()
        return body if isinstance(body, dict) else {}
    except Exception:
        return {}


@app.get("/api/lecturer/session")
async def lecturer_session(request: Request):
    lecturer = await current_lecturer_from_request(request)
    return {"authenticated": bool(lecturer), "lecturer": public_lecturer_view(lecturer) if lecturer else None}


@app.post("/api/lecturer/signup")
@limiter.limit("5/minute")
async def lecturer_signup(payload: dict[str, Any], request: Request):
    try:
        validated = LecturerSignupPayload.model_validate(payload)
        check_signup_allowed(validated.email, validated.inviteCode)
        lecturer = await repo.create_lecturer(validated)
        response = JSONResponse({"ok": True, "lecturer": public_lecturer_view(lecturer)})
        set_session_cookie(response, lecturer["id"], request)
        return response
    except ValidationError as exc:
        raise HTTPException(status_code=422, detail=validation_detail(exc)) from exc
    except PermissionError as exc:
        raise HTTPException(status_code=403, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    except SupabaseUnavailable as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.post("/api/lecturer/login")
@limiter.limit("10/minute")
async def lecturer_login(payload: dict[str, Any], request: Request):
    try:
        validated = LecturerLoginPayload.model_validate(payload)
        lecturer = await repo.get_lecturer_by_email(validated.email)
        if not lecturer:
            if repo.supabase_unavailable():
                raise HTTPException(status_code=503, detail="Supabase is unavailable. Please try again once it is restored.")
            raise HTTPException(status_code=401, detail="Incorrect email or password")
        if not await verify_password_async(validated.password, lecturer.get("password_hash", "")):
            raise HTTPException(status_code=401, detail="Incorrect email or password")
        response = JSONResponse({"ok": True, "lecturer": public_lecturer_view(lecturer)})
        set_session_cookie(response, lecturer["id"], request)
        return response
    except HTTPException:
        raise
    except ValidationError as exc:
        raise HTTPException(status_code=422, detail=validation_detail(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.post("/api/lecturer/logout")
def lecturer_logout():
    response = JSONResponse({"ok": True})
    clear_session_cookie(response)
    return response


@app.get("/api/subjects")
async def get_subjects():
    # Iterate a snapshot: a concurrent DELETE /api/subjects pops from SUBJECTS
    # while this coroutine is suspended on an await, which used to raise
    # "RuntimeError: dictionary changed size during iteration".
    subjects = list(SUBJECTS.items())
    # One query for every subject's counts instead of one round trip per
    # subject (the old code called list_tests() inside the loop).
    counts = await repo.get_test_counts()
    result = []
    for code, info in subjects:
        if code not in SUBJECTS:
            continue          # deleted while we were awaiting the counts query
        stats = counts.get(code, {"tests": 0, "questions": 0})
        builtin_questions = len(info.get("questions", []))
        result.append({
            "code": code,
            "name": info["name"],
            "questionCount": stats["questions"] or builtin_questions,
            "testCount": stats["tests"],
        })
    result.sort(key=lambda item: item["name"].lower())
    return result


@app.post("/api/session-token/{subject_code}")
async def create_session_token_endpoint(subject_code: str, request: Request):
    await require_lecturer(request)
    normalized = (subject_code or "").strip().upper()
    if normalized not in SUBJECTS:
        raise HTTPException(status_code=404, detail="Subject not found")
    token = generate_session_token(normalized)
    room = rooms.get(normalized)
    if room:
        room.current_token = token
    return {"ok": True, "token": token, "subject_code": normalized}


@app.get("/api/session-token/{token}/validate")
async def validate_session_token_endpoint(token: str):
    subject_code = lookup_session_token(token)
    if not subject_code:
        raise HTTPException(status_code=404, detail="This session link has expired or is invalid. Ask your lecturer for the current QR code.")
    subject = SUBJECTS.get(subject_code)
    return {
        "valid": True,
        "subject_code": subject_code,
        "subject_name": subject["name"] if subject else subject_code
    }


@app.post("/api/subjects")
async def create_subject(payload: dict[str, Any], request: Request):
    lecturer = await require_lecturer(request)
    try:
        validated = SubjectPayload.model_validate(payload)
        code = validated.code
        if code in SUBJECTS:
            raise HTTPException(status_code=409, detail="Subject code already exists.")
        created = await repo.create_subject(code, validated.name, lecturer)
        register_subject_in_catalog(code, created.get("name") or validated.name)
        return {"ok": True, "subject": {"code": code, "name": created.get("name") or validated.name}}
    except HTTPException:
        raise
    except ValidationError as exc:
        raise HTTPException(status_code=422, detail=validation_detail(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    except SupabaseUnavailable as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.delete("/api/subjects/{code}")
async def delete_subject(code: str, request: Request):
    lecturer = await require_lecturer(request)
    normalized = (code or "").strip().upper()
    if normalized in BUILTIN_SUBJECT_CODES:
        raise HTTPException(status_code=403, detail="Built-in subjects cannot be deleted.")
    if normalized not in SUBJECTS:
        raise HTTPException(status_code=404, detail="Subject not found")
    try:
        await repo.delete_subject(normalized, lecturer)
        for token, entry in list(session_tokens.items()):
            if (entry.get("subject_code") or "").strip().upper() == normalized:
                consume_session_token(token)
        SUBJECTS.pop(normalized, None)
        rooms.pop(normalized, None)
        return {"ok": True}
    except KeyError:
        raise HTTPException(status_code=404, detail="Subject not found") from None
    except PermissionError as exc:
        raise HTTPException(status_code=403, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    except SupabaseUnavailable as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.get("/api/tests/{subject_code}")
async def get_tests(subject_code: str, request: Request):
    if subject_code not in SUBJECTS:
        raise HTTPException(status_code=404, detail="Subject not found")
    lecturer = await require_lecturer(request)
    try:
        return await repo.list_tests(subject_code, lecturer.get("id"))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.get("/api/tests/{subject_code}/{test_id}")
async def get_test_detail(subject_code: str, test_id: str, request: Request):
    if subject_code not in SUBJECTS:
        raise HTTPException(status_code=404, detail="Subject not found")
    try:
        test_id = safe_test_id(test_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    lecturer = await require_lecturer(request)
    try:
        row = await repo.get_test(subject_code, test_id, lecturer.get("id"))
        if not row:
            raise HTTPException(status_code=404, detail="Test not found")
        return {
            "id": row["id"],
            "subject_code": row["subject_code"],
            "title": row.get("title") or "",
            "chapter": row.get("chapter") or "",
            "description": row.get("description") or "",
            "questions": row.get("questions") or [],
            "defaultTimeLimit": coerce_time_limit(row.get("default_time_limit")) or TIME_PER_Q,
            "questionCount": row.get("question_count") or len(row.get("questions") or []),
            "source": row.get("source", "supabase"),
            "ownerName": row.get("owner_name") or "System",
            "canEdit": bool(row.get("can_edit") or (row.get("created_by") == lecturer.get("id"))),
        }
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.post("/api/tests/{subject_code}")
@limiter.limit("30/minute")
async def create_test(subject_code: str, payload: dict[str, Any], request: Request):
    if subject_code not in SUBJECTS:
        raise HTTPException(status_code=404, detail="Subject not found")
    lecturer = await require_lecturer(request)
    try:
        validated = TestPayload.model_validate(payload)
        created = await repo.create_test(subject_code, validated, lecturer)
        try:
            await repo.clear_draft(subject_code, lecturer)
        except Exception:
            pass  # Draft clear is non-critical — never let it mask a successful test save
        return {"ok": True, "test": repo._summary(created, lecturer.get("id"))}
    except SupabaseUnavailable as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    except ValidationError as exc:
        raise HTTPException(status_code=422, detail=validation_detail(exc)) from exc
    except PermissionError as exc:
        raise HTTPException(status_code=403, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.put("/api/tests/{subject_code}/{test_id}")
@limiter.limit("60/minute")
async def update_test(subject_code: str, test_id: str, payload: dict[str, Any], request: Request):
    if subject_code not in SUBJECTS:
        raise HTTPException(status_code=404, detail="Subject not found")
    try:
        test_id = safe_test_id(test_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    lecturer = await require_lecturer(request)
    try:
        validated = TestPayload.model_validate(payload)
        updated = await repo.update_test(subject_code, test_id, validated, lecturer)
        try:
            await repo.clear_draft(subject_code, lecturer)
        except Exception:
            pass  # Draft clear is non-critical — never let it mask a successful test update
        return {"ok": True, "test": repo._summary(updated, lecturer.get("id"))}
    except SupabaseUnavailable as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    except ValidationError as exc:
        raise HTTPException(status_code=422, detail=validation_detail(exc)) from exc
    except KeyError:
        raise HTTPException(status_code=404, detail="Test not found") from None
    except PermissionError as exc:
        raise HTTPException(status_code=403, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.delete("/api/tests/{subject_code}/{test_id}")
async def delete_test(subject_code: str, test_id: str, request: Request):
    if subject_code not in SUBJECTS:
        raise HTTPException(status_code=404, detail="Subject not found")
    try:
        test_id = safe_test_id(test_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    lecturer = await require_lecturer(request)
    try:
        await repo.delete_test(subject_code, test_id, lecturer)
        return {"ok": True}
    except SupabaseUnavailable as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    except KeyError:
        raise HTTPException(status_code=404, detail="Test not found") from None
    except PermissionError as exc:
        raise HTTPException(status_code=403, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


DRAFT_BACKEND_LABELS = {
    "supabase": "Supabase",
    "local-file": "this server's local file — it will be lost if the server redeploys",
    "memory": "server memory only — it will be lost when the server restarts",
}


@app.get("/api/drafts/{subject_code}")
async def get_test_draft(subject_code: str, request: Request):
    if subject_code not in SUBJECTS:
        raise HTTPException(status_code=404, detail="Subject not found")
    lecturer = await require_lecturer(request)
    try:
        draft, backend, error = await repo.get_draft(subject_code, lecturer)
    except Exception as exc:
        # A missing quiz_test_drafts table must not break the editor, but it must
        # not be invisible either — the client shows this on the draft status line.
        return {"draft": None, "storedIn": None, "error": str(exc)}
    return {"draft": draft, "storedIn": backend if draft else None, "error": error}


@app.post("/api/drafts/{subject_code}")
@limiter.limit(DRAFT_RATE_LIMIT)
async def save_test_draft(subject_code: str, payload: dict[str, Any], request: Request):
    if subject_code not in SUBJECTS:
        raise HTTPException(status_code=404, detail="Subject not found")
    lecturer = await require_lecturer(request)
    try:
        validated = DraftPayload.model_validate(payload)
    except ValidationError as exc:
        raise HTTPException(status_code=422, detail=validation_detail(exc)) from exc
    try:
        draft, backend, remote_error = await repo.save_draft(subject_code, lecturer, validated)
    except SupabaseUnavailable as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    except Exception as exc:
        # The draft genuinely was not written anywhere. Say so with a real error
        # status: the old code returned HTTP 200 with {"ok": false} and the
        # editor happily reported "Draft saved".
        raise HTTPException(status_code=500, detail=f"Draft could not be saved: {exc}") from exc
    return {
        "ok": True,
        "draft": draft,
        "storedIn": backend,
        "storedInLabel": DRAFT_BACKEND_LABELS.get(backend, backend),
        "degraded": backend != "supabase" and repo.supabase_configured,
        "error": remote_error,
    }


@app.get("/api/diagnostics")
async def diagnostics(request: Request):
    """Probe each storage table so silent breakage is diagnosable.

    Lecturer-authenticated: the probe reports Supabase error text, which is not
    something anonymous callers should see.
    """
    await require_lecturer(request)
    tables = {
        "quiz_lecturers": "lecturers_base",
        "quiz_tests": "quiz_tests_base",
        "quiz_test_drafts": "drafts_base",
        "quiz_subjects": "subjects_base",
    }
    results: dict[str, Any] = {}
    remote = repo.remote
    for table, attr in tables.items():
        if remote is None:
            results[table] = {
                "reachable": False,
                "error": "Supabase is not configured on this server." if not repo.supabase_configured
                         else (repo.supabase_error or "Supabase connection is disabled."),
            }
            continue
        try:
            await remote._request("GET", getattr(remote, attr), params={"select": "*", "limit": "1"})
            results[table] = {"reachable": True, "error": None}
        except Exception as exc:
            results[table] = {"reachable": False, "error": str(exc)}
    return {
        "storage": repo.get_storage_status(),
        "draftError": repo.draft_error,
        "localStore": {
            "enabled": repo.local_store_enabled,
            "path": str(repo.local_store_path),
            "exists": repo.local_store_path.exists(),
            "error": repo.local_store_error,
        },
        "tables": results,
    }


@app.delete("/api/drafts/{subject_code}")
async def clear_test_draft(subject_code: str, request: Request):
    if subject_code not in SUBJECTS:
        raise HTTPException(status_code=404, detail="Subject not found")
    lecturer = await require_lecturer(request)
    try:
        await repo.clear_draft(subject_code, lecturer)
        return {"ok": True}
    except Exception:
        # Always return ok — draft clearing is never worth surfacing as an error.
        return {"ok": True}


@app.post("/api/import/questions")
@limiter.limit("20/minute")
async def import_questions(request: Request, file: UploadFile = File(...)):
    """Parse a .docx into reviewable questions.

    Returns JSON only. It never writes a test — the lecturer reviews the parsed
    questions in the editor and saves them as normal.
    """
    await require_lecturer(request)
    filename = (file.filename or "").strip()
    if not filename.lower().endswith(".docx"):
        raise HTTPException(
            status_code=400,
            detail="Please upload a .docx file. Word's older .doc format and PDFs cannot be read — "
                   "open the file in Word and use File > Save As > Word Document (.docx).",
        )
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="That file is empty.")
    if len(data) > docx_import.MAX_UPLOAD_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"That file is larger than {docx_import.MAX_UPLOAD_BYTES // (1024 * 1024)} MB.",
        )
    try:
        result = await asyncio.to_thread(docx_import.parse_docx, data)
    except docx_import.DocxImportError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Could not read that document: {exc}") from exc
    return result


@app.get("/api/import/template")
async def import_template(request: Request):
    """Generated at request time, so no binary lives in the repository."""
    await require_lecturer(request)
    try:
        data = await asyncio.to_thread(docx_import.build_template_docx)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Could not build the template: {exc}") from exc
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": 'attachment; filename="quiz_question_template.docx"'},
    )


@app.get("/api/results/{subject_code}")
async def list_stored_results(subject_code: str, request: Request):
    """Stored sessions for a subject, plus what they cost in storage.

    Personal data (names, student numbers, answers) is deliberately excluded
    from this listing; it is only in the individual result.
    """
    await require_lecturer(request)
    try:
        subject_code = safe_subject_code(subject_code)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if subject_code not in SUBJECTS:
        raise HTTPException(status_code=404, detail="Subject not found")
    results = await repo.list_game_results(subject_code)
    return {
        "results": results,
        "storage": repo.results_storage_usage(subject_code),
    }


@app.get("/api/results/{subject_code}/{result_id}")
async def get_stored_result(subject_code: str, result_id: str, request: Request):
    await require_lecturer(request)
    try:
        subject_code = safe_subject_code(subject_code)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    row = await repo.get_game_result(subject_code, result_id)
    if not row:
        raise HTTPException(status_code=404, detail="No stored result with that id.")
    return row


@app.get("/api/export/tests")
async def export_tests(request: Request):
    lecturer = await require_lecturer(request)
    try:
        tests = await repo.list_tests_by_creator(lecturer["id"])
        stamp = datetime.now(UTC).strftime("%Y%m%d")
        filename = f"quiz_backup_{stamp}.json"
        payload = json.dumps(tests, ensure_ascii=False, indent=2)
        headers = {"Content-Disposition": f"attachment; filename={filename}"}
        return Response(content=payload, media_type="application/json", headers=headers)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.get("/api/stats/{subject_code}")
async def download_stats(subject_code: str, request: Request):
    # This workbook contains student names, student numbers and per-question
    # results. It was downloadable by anyone who knew a subject code.
    # The host browser is already signed in, so the end-of-game automatic
    # download still works — it sends the session cookie same-origin.
    await require_lecturer(request)
    if subject_code not in rooms:
        raise HTTPException(status_code=404, detail="Subject not found")

    room = rooms[subject_code]
    stats = room.last_game_stats
    if not stats or not stats["players"]:
        raise HTTPException(status_code=404, detail="No game data available. Play a game first.")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed") from None

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Student Results"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    correct_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    wrong_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    no_answer_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    questions = stats["questions"]
    num_q = len(questions)
    headers = ["Rank", "Student Name", "Student Number", "Total Score"]
    for i in range(num_q):
        headers.append(f"Q{i+1}")
        headers.append(f"Q{i+1} Time (s)")
    headers.extend(["Questions Correct", "Accuracy %"])

    for col, heading in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=heading)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    sorted_players = sorted(stats["players"].values(), key=lambda p: -p["score"])
    for rank, player in enumerate(sorted_players, 1):
        row = rank + 1
        ws1.cell(row=row, column=1, value=rank).border = thin_border
        ws1.cell(row=row, column=2, value=player["name"]).border = thin_border
        ws1.cell(row=row, column=3, value=player.get("student_number", "")).border = thin_border
        ws1.cell(row=row, column=4, value=player["score"]).border = thin_border
        correct_count = 0
        col_offset = 5
        for qi in range(num_q):
            result_col = col_offset + qi * 2
            time_col = result_col + 1
            if qi < len(player["answers"]):
                ans = player["answers"][qi]
                is_correct = ans.get("correct", False)
                points = ans.get("points", 0)
                time_taken = round(ans.get("time", 0), 1) if ans.get("time") else "-"
                if is_correct:
                    correct_count += 1
                result_cell = ws1.cell(row=row, column=result_col)
                if ans.get("choice", -1) == -1:
                    result_cell.value = "No answer"
                    result_cell.fill = no_answer_fill
                elif is_correct:
                    result_cell.value = f"✓ (+{points})"
                    result_cell.fill = correct_fill
                else:
                    chosen = ans.get("choice", -1)
                    if 0 <= chosen < len(questions[qi]["options"]):
                        result_cell.value = f"✗ ({questions[qi]['options'][chosen][:20]})"
                    else:
                        result_cell.value = "✗"
                    result_cell.fill = wrong_fill
                result_cell.border = thin_border
                result_cell.alignment = Alignment(horizontal="center")
                time_cell = ws1.cell(row=row, column=time_col, value=time_taken)
                time_cell.border = thin_border
                time_cell.alignment = Alignment(horizontal="center")
            else:
                ws1.cell(row=row, column=result_col, value="-").border = thin_border
                ws1.cell(row=row, column=time_col, value="-").border = thin_border
        summary_col = col_offset + num_q * 2
        ws1.cell(row=row, column=summary_col, value=f"{correct_count}/{num_q}").border = thin_border
        accuracy = round((correct_count / num_q) * 100, 1) if num_q else 0
        ws1.cell(row=row, column=summary_col + 1, value=f"{accuracy}%").border = thin_border

    for col in ws1.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws1.column_dimensions[col_letter].width = min(max_len + 3, 28)

    ws2 = wb.create_sheet("Question Analysis")
    q_headers = ["Question #", "Question Text", "Correct Answer", "# Correct", "# Wrong", "# No Answer", "% Correct", "Avg Time (s)"]
    for col, heading in enumerate(q_headers, 1):
        cell = ws2.cell(row=1, column=col, value=heading)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    for qi, question in enumerate(questions):
        row = qi + 2
        correct_count = wrong_count = no_answer_count = 0
        total_time = 0
        time_count = 0
        for player in stats["players"].values():
            if qi < len(player["answers"]):
                ans = player["answers"][qi]
                if ans.get("choice", -1) == -1:
                    no_answer_count += 1
                elif ans.get("correct"):
                    correct_count += 1
                    if ans.get("time"):
                        total_time += ans["time"]
                        time_count += 1
                else:
                    wrong_count += 1
                    if ans.get("time"):
                        total_time += ans["time"]
                        time_count += 1
        total_answered = correct_count + wrong_count + no_answer_count
        pct_correct = round((correct_count / total_answered) * 100, 1) if total_answered else 0
        avg_time = round(total_time / time_count, 1) if time_count else "-"
        ws2.cell(row=row, column=1, value=qi + 1).border = thin_border
        ws2.cell(row=row, column=2, value=question["q"][:100]).border = thin_border
        ws2.cell(row=row, column=3, value=question["options"][question["correct"]]).border = thin_border
        ws2.cell(row=row, column=4, value=correct_count).border = thin_border
        ws2.cell(row=row, column=5, value=wrong_count).border = thin_border
        ws2.cell(row=row, column=6, value=no_answer_count).border = thin_border
        pct_cell = ws2.cell(row=row, column=7, value=f"{pct_correct}%")
        pct_cell.border = thin_border
        if pct_correct < 50:
            pct_cell.fill = wrong_fill
        elif pct_correct >= 80:
            pct_cell.fill = correct_fill
        ws2.cell(row=row, column=8, value=avg_time).border = thin_border

    for col in ws2.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws2.column_dimensions[col_letter].width = min(max_len + 3, 32)

    summary = wb.create_sheet("Game Summary")
    summary["A1"] = "Subject"
    summary["B1"] = stats["subject_name"]
    summary["A2"] = "Subject Code"
    summary["B2"] = stats["subject_code"]
    summary["A3"] = "Test Title"
    summary["B3"] = stats.get("test_title") or ""
    summary["A4"] = "Chapter"
    summary["B4"] = stats.get("test_chapter") or ""
    summary["A5"] = "Played At"
    summary["B5"] = stats["timestamp"]
    summary["A6"] = "Players"
    summary["B6"] = len(stats["players"])
    summary["A7"] = "Questions"
    summary["B7"] = len(stats["questions"])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    def _safe_filename_part(value: str) -> str:
        """Strip non-ASCII and filesystem-unsafe characters, replace spaces with underscores."""
        import unicodedata
        # Normalise accented characters to their ASCII base where possible
        value = unicodedata.normalize("NFKD", value)
        # Keep only printable ASCII, replace spaces, drop the rest
        result = []
        for ch in value:
            if ch == " ":
                result.append("_")
            elif ch in r'/\:*?"<>|':
                result.append("-")
            elif 0x20 <= ord(ch) <= 0x7E:
                result.append(ch)
            # Non-ASCII characters (e.g. em dash —) are silently dropped
        return "".join(result)[:60].strip("_-") or "Stats"

    safe_title = _safe_filename_part(stats.get("test_title") or subject_code)
    safe_session = _safe_filename_part(room.last_game_stats.get("session_name") or safe_title)
    filename = f"Stats_{subject_code}_{safe_session}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ──────────────────────────────────────────────────────────────────────────────
# WebSocket helpers
# ──────────────────────────────────────────────────────────────────────────────
def get_active_test_meta(room: GameRoom) -> dict[str, Any] | None:
    if not room.active_test_id:
        return None
    return {
        "id": room.active_test_id,
        "title": room.active_test_title,
        "chapter": room.active_test_chapter,
        "questionCount": room.total_q,
        "defaultTimeLimit": room.default_time_limit,
        "estimatedSeconds": estimate_test_seconds({
            "default_time_limit": room.default_time_limit,
            "questions": room.questions,
        }),
    }


def is_participating_player(room: GameRoom, player: dict[str, Any] | None) -> bool:
    if not room.game_code_enabled:
        return True
    return bool(player and player.get("game_code_verified"))


def student_identity_key(student_number: str, name: str) -> str:
    """A stable key for one student within a session.

    Matching used to be an exact string comparison of the student number, so
    "221 012 345" and "221012345" looked like two different people and a
    returning student was treated as a stranger. Spaces, dashes and case are
    ignored here.
    """
    number = re.sub(r"[^0-9a-z]", "", (student_number or "").lower())
    if number:
        return f"number:{number}"
    return "name:" + re.sub(r"\s+", " ", (name or "").strip().lower())


def find_existing_player(
    room: GameRoom,
    *,
    visitor_id: str,
    name: str,
    student_number: str
) -> tuple[str | None, dict[str, Any] | None]:
    wanted = student_identity_key(student_number, name)
    for vid, player in room.players.items():
        if vid == visitor_id:
            continue
        if student_identity_key(player.get("student_number", ""), player.get("name", "")) == wanted:
            return vid, player
    return None, None


def build_joined_payload(room: GameRoom, visitor_id: str) -> dict[str, Any]:
    joined_payload = {
        "type": "joined",
        "phase": room.phase,
        "playerId": visitor_id,
        "playerCount": len(room.players),
        "activeTest": get_active_test_meta(room)
    }
    if room.phase == "question":
        q = room.questions[room.current_q]
        limit = room.time_limit_for()
        remaining = max(0, limit - room.question_elapsed)
        joined_payload["currentQuestion"] = {
            "question": q["q"],
            "options": q["options"],
            "qNum": room.current_q + 1,
            "totalQ": room.total_q,
            "timeLimit": limit,
            "remaining": round(remaining, 2)
        }
        joined_payload["alreadyAnswered"] = visitor_id in room.answers_this_round
    elif room.phase in ("reveal", "get_ready", "final"):
        joined_payload["phase"] = room.phase
    return joined_payload


def get_player_list(room: GameRoom, *, participant_only: bool = False) -> list[dict[str, Any]]:
    players = []
    for vid, p in room.players.items():
        if participant_only and not is_participating_player(room, p):
            continue
        players.append({
            "id": vid,
            "name": p["name"],
            "score": p["score"],
            "connected": p.get("ws") is not None
        })
    players.sort(key=lambda x: -x["score"])
    return players


def get_leaderboard(room: GameRoom, *, participant_only: bool = False) -> list[dict[str, Any]]:
    players = get_player_list(room, participant_only=participant_only)
    for i, p in enumerate(players):
        p["rank"] = i + 1
    return players


async def broadcast_to_players(room: GameRoom, msg: dict[str, Any] | str, *, participant_only: bool = False) -> None:
    text = msg if isinstance(msg, str) else json.dumps(msg)

    async def _safe_send(ws: WebSocket):
        try:
            await ws.send_text(text)
        except Exception:
            pass

    tasks = []
    for _, player in list(room.players.items()):
        if participant_only and not is_participating_player(room, player):
            continue
        ws = player.get("ws")
        if ws is not None:
            tasks.append(_safe_send(ws))
    if tasks:
        await asyncio.gather(*tasks, return_exceptions=True)


async def broadcast_to_selected_players(room: GameRoom, msg: dict[str, Any] | str, player_ids: set[str]) -> None:
    text = msg if isinstance(msg, str) else json.dumps(msg)

    async def _safe_send(ws: WebSocket):
        try:
            await ws.send_text(text)
        except Exception:
            pass

    tasks = []
    for vid, player in list(room.players.items()):
        if vid not in player_ids:
            continue
        ws = player.get("ws")
        if ws is not None:
            tasks.append(_safe_send(ws))
    if tasks:
        await asyncio.gather(*tasks, return_exceptions=True)


async def send_to_host(room: GameRoom, msg: dict[str, Any]) -> None:
    if room.host_ws is None:
        return
    try:
        await room.host_ws.send_text(json.dumps(msg))
    except Exception:
        room.host_ws = None


async def _do_push_room_update(room: GameRoom) -> None:
    host_payload = {
        "type": "player_update",
        "players": get_player_list(room),
        "activeTest": get_active_test_meta(room)
    }
    player_payload = json.dumps({
        "type": "player_update",
        "playerCount": len(room.players)
    })
    await send_to_host(room, host_payload)
    await broadcast_to_players(room, player_payload)


async def push_room_update(room: GameRoom) -> None:
    # Skip debounce during active game phases so answer counts stay responsive.
    if room.phase in ("question", "reveal"):
        await _do_push_room_update(room)
        return

    # Cancel any already-pending debounced broadcast and schedule a fresh one.
    if room._pending_room_update is not None:
        room._pending_room_update.cancel()

    async def _debounced():
        await asyncio.sleep(0.3)
        room._pending_room_update = None
        await _do_push_room_update(room)

    room._pending_room_update = asyncio.create_task(_debounced())


def mark_unanswered_players(room: GameRoom) -> None:
    for vid in room.players:
        if not is_participating_player(room, room.players.get(vid)):
            continue
        if vid not in room.answers_this_round:
            room.players[vid]["streak"] = 0
            room.players[vid]["answers"].append({
                "q": room.current_q,
                "choice": -1,
                "correct": False,
                "points": 0,
                "time": 0
            })


async def sync_answer_count(room: GameRoom) -> None:
    if room.phase != "question":
        return
    answered_count = len(room.answers_this_round)
    total_connected = sum(
        1
        for _, player in room.players.items()
        if player.get("ws") is not None and is_participating_player(room, player)
    )
    await send_to_host(room, {
        "type": "answer_count",
        "answered": answered_count,
        "total": total_connected
    })


async def maybe_finish_question_early(room: GameRoom) -> None:
    # The lock makes the phase check and the phase change atomic across
    # concurrent answer handlers and the question timer, so
    # mark_unanswered_players() can never run twice for one question and append
    # duplicate answer records.
    async with room.advance_lock:
        if room.phase != "question":
            return
        answered_count = len(room.answers_this_round)
        total_connected = sum(
            1
            for _, player in room.players.items()
            if player.get("ws") is not None and is_participating_player(room, player)
        )
        if answered_count < total_connected or total_connected <= 0:
            return
        if room.question_timer_task and not room.question_timer_task.done():
            room.question_timer_task.cancel()
        mark_unanswered_players(room)
        room.phase = "reveal"
    await auto_reveal(room)


async def kick_player_from_room(room: GameRoom, player_id: str, *, message: str) -> bool:
    player = room.players.get(player_id)
    if not player:
        return False
    room.answers_this_round.pop(player_id, None)
    ws = player.get("ws")
    # Keep player record so they can rejoin; only disconnect the WebSocket.
    player["ws"] = None
    if ws is not None:
        try:
            await ws.send_text(json.dumps({
                "type": "kicked",
                "message": message
            }))
        except Exception:
            pass
        try:
            await ws.close(code=4002)
        except Exception:
            pass
    await push_room_update(room)
    await sync_answer_count(room)
    await maybe_finish_question_early(room)
    return True


async def cancel_question_timer(room: GameRoom) -> None:
    current = asyncio.current_task()
    for attr in ("question_timer_task", "start_task"):
        task = getattr(room, attr, None)
        setattr(room, attr, None)
        if task is None or task.done():
            continue
        if task is current:
            # We are running *inside* that task — reached here via
            # _timer -> auto_reveal -> advance_to_next -> force_end_game.
            # Cancelling would abort the end-of-game work we are doing, and
            # awaiting it raises "Task cannot await on itself". Just detach.
            continue
        task.cancel()
        try:
            await task
        except asyncio.CancelledError:
            pass
        except Exception:
            pass


async def return_room_to_lobby(room: GameRoom, *, keep_players: bool) -> None:
    # Generate a fresh session token (this also consumes the previous one)
    new_token = generate_session_token(room.subject_code)
    room.current_token = new_token
    room.game_code = ""
    room.game_code_enabled = False
    await cancel_question_timer(room)
    room.phase = "lobby"
    room.current_q = 0
    room.question_start_time = 0
    room.answers_this_round = {}
    room.paused = False

    if keep_players:
        for player in list(room.players.values()) + list(room.session_roster.values()):
            player["score"] = 0
            player["streak"] = 0
            player["answers"] = []
        await broadcast_to_players(room, {
            "type": "reset",
            "phase": "lobby",
            "playerCount": len(room.players),
            "activeTest": get_active_test_meta(room)
        })
    else:
        await broadcast_to_players(room, {
            "type": "reset",
            "phase": "lobby",
            "playerCount": 0,
            "activeTest": get_active_test_meta(room)
        })
        room.players = {}
        room.session_roster = {}

    await send_to_host(room, {
        "type": "host_joined",
        "phase": "lobby",
        "players": get_player_list(room),
        "currentQ": 0,
        "totalQ": room.total_q,
        "subjectCode": room.subject_code,
        "subjectName": room.subject_name,
        "selectedTest": get_active_test_meta(room),
        "gameCode": room.game_code,
        "gameCodeEnabled": room.game_code_enabled,
        "hasQuestions": room.total_q > 0,
        "hasStats": room.last_game_stats is not None,
        "sessionToken": new_token
    })
    await push_room_update(room)


def build_answer_distribution(room: GameRoom) -> list[int]:
    """How many students picked each option this round.

    The server already holds every choice in answers_this_round, so this is
    nearly free — and it is the most useful teaching signal in a live quiz.
    """
    question = room.questions[room.current_q] if 0 <= room.current_q < len(room.questions) else None
    option_count = len(question.get("options", [])) if isinstance(question, dict) else 4
    counts = [0] * option_count
    for vid, answer in room.answers_this_round.items():
        if not is_participating_player(room, room.players.get(vid)):
            continue
        choice = answer.get("choice", -1)
        if isinstance(choice, int) and 0 <= choice < option_count:
            counts[choice] += 1
    return counts


def build_student_review(room: GameRoom) -> dict[str, list[dict[str, Any]]]:
    """Per-student list of what they got wrong, with the correct answer."""
    review: dict[str, list[dict[str, Any]]] = {}
    for vid, player in room.players.items():
        if not is_participating_player(room, player):
            continue
        entries = []
        for answer in player.get("answers") or []:
            index = answer.get("q")
            if not isinstance(index, int) or not 0 <= index < len(room.questions):
                continue
            question = room.questions[index]
            choice = answer.get("choice", -1)
            options = question.get("options", [])
            entries.append({
                "qNum": index + 1,
                "question": question.get("q", ""),
                "options": options,
                "yourChoice": choice if isinstance(choice, int) else -1,
                "correct": question.get("correct", 0),
                "wasCorrect": bool(answer.get("correct")),
                "points": answer.get("points", 0),
                "explanation": question.get("explanation", ""),
            })
        review[vid] = entries
    return review


async def force_end_game(room: GameRoom) -> None:
    lb = get_leaderboard(room, participant_only=True)
    participant_ids = {
        vid for vid, player in room.players.items()
        if is_participating_player(room, player)
    }
    non_participant_ids = set(room.players.keys()) - participant_ids
    active_token = get_room_active_token(room)
    if active_token:
        consume_session_token(active_token)
    room.game_code = ""
    room.game_code_enabled = False
    await cancel_question_timer(room)
    room.phase = "final"
    room.archive_stats()
    # Store a lightweight copy so a redeploy cannot destroy results that were
    # never downloaded. Non-fatal by design: the automatic download is the
    # record that matters, and this must never break the end of a game.
    if PERSIST_RESULTS and room.last_game_stats:
        try:
            await repo.store_game_result(room.last_game_stats, room.host_lecturer_id)
        except Exception as exc:
            print(f"[results] could not store results for {room.subject_code}: {exc}")
    # Per-question answer distribution, so students can see how the class did.
    review = build_student_review(room)
    await broadcast_to_selected_players(room, {"type": "final", "leaderboard": lb}, participant_ids)
    await broadcast_to_selected_players(room, {"type": "game_ended"}, non_participant_ids)
    await send_to_host(room, {"type": "final", "leaderboard": lb, "hasStats": True})
    # Each student gets only their own answers, alongside the correct answer and
    # the explanation.
    for vid in participant_ids:
        player = room.players.get(vid)
        ws = player.get("ws") if player else None
        if ws is None:
            continue
        try:
            await ws.send_text(json.dumps({"type": "review", "questions": review.get(vid, [])}))
        except Exception:
            pass
    room.players = {}
    room.session_roster = {}


# ──────────────────────────────────────────────────────────────────────────────
# WebSocket endpoint
# ──────────────────────────────────────────────────────────────────────────────
@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await websocket.accept()
    # Identity comes from a server-signed token only. A bare visitorId query
    # parameter is no longer trusted: it let a student claim another student's
    # id and inherit their score.
    visitor_id = parse_visitor_token(websocket.query_params.get("vt")) or str(uuid.uuid4())
    role = None
    room = None

    try:
        while True:
            raw = await websocket.receive_text()
            # Cheap resource-exhaustion guard: a single socket could otherwise
            # push arbitrarily large frames at the server.
            if len(raw) > MAX_WS_MESSAGE_BYTES:
                await websocket.send_text(json.dumps({"type": "error", "message": "Message too large."}))
                await websocket.close(code=1009)
                break
            try:
                msg = json.loads(raw)
            except (json.JSONDecodeError, ValueError):
                continue
            if not isinstance(msg, dict):
                continue
            action = msg.get("action")

            if action == "host_join":
                lecturer = await current_lecturer_from_websocket(websocket)
                if not lecturer:
                    await websocket.send_text(json.dumps({"type": "auth_required", "message": "Lecturer sign-in required"}))
                    continue
                subject_code = msg.get("subject")
                test_id = msg.get("testId")
                host_token = (msg.get("token") or "").strip().upper()
                if subject_code not in rooms:
                    await websocket.send_text(json.dumps({"type": "error", "message": "Invalid subject"}))
                    continue
                test_data = await repo.get_test(subject_code, test_id) if test_id else None
                if not test_data:
                    await websocket.send_text(json.dumps({"type": "error", "message": "Test not found"}))
                    continue

                role = "host"
                room = rooms[subject_code]

                # A room belongs to a subject, not to a test, and lives in
                # process memory — so it keeps whatever phase the previous
                # session left it in. When a game finishes the room sits in
                # "final" indefinitely, and every student scanning the QR code
                # is told "the game is already in progress" while the host sees
                # a perfectly normal-looking lobby.
                #
                # This used to reset only when a *different test* was chosen, so
                # picking up the same test as the previous session — very common
                # when another lecturer re-runs a colleague's quiz — left the
                # room stuck and locked the whole class out.
                requested_new_test = bool(test_id and test_id != room.active_test_id)
                different_lecturer = bool(room.host_lecturer_id and room.host_lecturer_id != lecturer.get("id"))
                explicit_new_session = bool(msg.get("newSession"))

                # A finished game is not "in progress": anything that looks like
                # a fresh start clears it.
                needs_reset = room.phase == "final" and (
                    requested_new_test or different_lecturer or explicit_new_session
                )
                # A room abandoned mid-game (host closed the laptop, or a paused
                # question that never resumed) is only cleared on a deliberate
                # new session, never on a plain reconnect — that would wipe a
                # live game out from under the class.
                if room.phase not in ("lobby", "final") and explicit_new_session:
                    needs_reset = True

                if room.phase == "lobby" or needs_reset:
                    room.set_active_test(test_data)
                room.session_name = msg.get("sessionName", "").strip()[:80] or room.active_test_title
                if host_token and lookup_session_token(host_token) == subject_code:
                    room.current_token = host_token
                elif room.phase == "lobby":
                    room.current_token = ""
                room.host_ws = websocket
                room.host_visitor = visitor_id
                room.host_lecturer_id = lecturer.get("id")

                if needs_reset:
                    # Players from the previous session are long gone; starting
                    # a new session should not inherit their records.
                    await return_room_to_lobby(room, keep_players=False)
                    continue

                await websocket.send_text(json.dumps({
                    "type": "host_joined",
                    "phase": room.phase,
                    "players": get_player_list(room),
                    "currentQ": room.current_q,
                    "totalQ": room.total_q,
                    "subjectCode": room.subject_code,
                    "subjectName": room.subject_name,
                    "selectedTest": get_active_test_meta(room),
                    "gameCode": room.game_code,
                    "gameCodeEnabled": room.game_code_enabled,
                    "hasQuestions": room.total_q > 0,
                    "hasStats": room.last_game_stats is not None
                }))
                await push_room_update(room)

            elif action == "start_game":
                if role != "host" or room is None:
                    continue
                if room.total_q == 0:
                    await websocket.send_text(json.dumps({"type": "error", "message": "No questions loaded for this test."}))
                    continue
                if room.start_task is not None and not room.start_task.done():
                    continue                       # already starting
                # Run the countdown in the background. It used to sit inside this
                # receive loop, so for 20 seconds the host could not pause or
                # cancel and every queued message stalled behind it.
                room.start_task = asyncio.create_task(run_game_start(room, shuffle=bool(msg.get("shuffle")), use_code=bool(msg.get("useCode"))))

            elif action == "next_question":
                if role == "host" and room is not None:
                    await advance_to_next(room, from_question=room.current_q)

            elif action == "extend_time":
                # Give the room longer on the question currently running. The
                # timer counts against room.question_time_limit, so raising it
                # extends the live question.
                if role != "host" or room is None or room.phase != "question":
                    continue
                extra = coerce_time_limit(msg.get("seconds")) or 15
                extra = max(5, min(60, extra))
                room.question_time_limit = min(MAX_TIME_LIMIT, room.question_time_limit + extra)
                remaining = max(0.0, room.question_time_limit - room.question_elapsed)
                payload = {
                    "type": "time_extended",
                    "addedSeconds": extra,
                    "timeLimit": room.question_time_limit,
                    "remaining": round(remaining, 2),
                }
                await send_to_host(room, payload)
                await broadcast_to_players(room, payload, participant_only=True)

            elif action == "host_pause":
                if role != "host" or room is None:
                    continue
                room.paused = not room.paused
                await send_to_host(room, {"type": "pause_state", "paused": room.paused})
                await broadcast_to_players(room, {
                    "type": "pause_state",
                    "paused": room.paused
                })

            elif action == "reset_game":
                if role == "host" and room is not None:
                    room.archive_stats()
                    await return_room_to_lobby(room, keep_players=True)

            elif action == "cancel_game":
                if role == "host" and room is not None:
                    await return_room_to_lobby(room, keep_players=True)

            elif action == "end_game":
                if role == "host" and room is not None:
                    await force_end_game(room)

            elif action == "kick_player":
                if role != "host" or room is None:
                    continue
                player_id = (msg.get("playerId") or "").strip()
                if not player_id:
                    continue
                await kick_player_from_room(
                    room,
                    player_id,
                    message="The lecturer removed you from this session."
                )

            elif action == "ping":
                await websocket.send_text(json.dumps({"type": "pong"}))

            elif action == "player_join":
                token = (msg.get("token") or "").strip().upper()
                subject_code_from_token = lookup_session_token(token) if token else None
                subject_code = subject_code_from_token or (msg.get("subject") or "").strip().upper()
                if subject_code not in rooms:
                    await websocket.send_text(json.dumps({
                        "type": "error",
                        "message": "This session link has expired or is invalid. Ask your lecturer for the current QR code."
                    }))
                    continue
                room = rooms[subject_code]
                name = msg.get("name", "Anonymous").strip()[:20]
                student_number = msg.get("studentNumber", "").strip()[:20]
                provided_code = (msg.get("gameCode") or "").strip()
                existing_vid, existing_player = find_existing_player(
                    room,
                    visitor_id=visitor_id,
                    name=name,
                    student_number=student_number
                )
                student_key = student_identity_key(student_number, name)
                # A student who has already taken part in this session is not a
                # new player, even if their live record has gone (they left, or
                # their browser identity changed after a reconnect). Without
                # this they were told "the game is already in progress" and had
                # no way back in.
                roster_player = None
                if existing_player is None and visitor_id not in room.players:
                    roster_player = room.session_roster.get(student_key)
                required_room_token = (room.current_token or "").strip().upper()
                is_known_player = (
                    visitor_id in room.players
                    or existing_player is not None
                    or roster_player is not None
                )
                # Reject on expired/invalid token only for players not already in the room.
                # Known players (already mid-game) must always be allowed to reconnect.
                if token and not subject_code_from_token and not is_known_player:
                    await websocket.send_text(json.dumps({
                        "type": "error",
                        "message": "This session link has expired. Ask your lecturer for the current QR code."
                    }))
                    continue
                if required_room_token and not subject_code_from_token and not is_known_player:
                    await websocket.send_text(json.dumps({
                        "type": "error",
                        "message": "This session link has expired or is invalid. Ask your lecturer for the current QR code."
                    }))
                    continue
                current_player = room.players.get(visitor_id)
                can_bypass_game_code = (
                    is_participating_player(room, existing_player)
                    or is_participating_player(room, current_player)
                )
                if room.game_code_enabled and not can_bypass_game_code and provided_code != room.game_code:
                    await websocket.send_text(json.dumps({
                        "type": "error_game_code",
                        "message": "Enter the 4-digit code shown on the lecturer screen to continue."
                    }))
                    continue
                if room.phase != "lobby" and not is_known_player:
                    await websocket.send_text(json.dumps({
                        "type": "error",
                        "message": "The game is already in progress. You cannot join as a new player at this stage."
                    }))
                    continue
                if not is_known_player and len(room.players) >= MAX_PLAYERS_PER_ROOM:
                    await websocket.send_text(json.dumps({
                        "type": "error",
                        "message": f"This session is full ({MAX_PLAYERS_PER_ROOM} students). Ask your lecturer to start another session."
                    }))
                    continue
                if subject_code_from_token and room.current_token != token:
                    room.current_token = token
                role = "player"
                if existing_player:
                    old_ws = existing_player.get("ws")
                    if old_ws is not None:
                        try:
                            await old_ws.close(code=4001)
                        except Exception:
                            pass
                    room.players.pop(existing_vid, None)
                    existing_player["name"] = name
                    existing_player["student_number"] = student_number
                    existing_player["ws"] = websocket
                    existing_player["game_code_verified"] = (
                        is_participating_player(room, existing_player)
                        or not room.game_code_enabled
                        or provided_code == room.game_code
                    )
                    room.players[visitor_id] = existing_player
                elif roster_player is not None:
                    # Returning after their record left room.players. Reusing the
                    # same record keeps their score and answers so far.
                    roster_player["name"] = name
                    roster_player["student_number"] = student_number
                    roster_player["ws"] = websocket
                    roster_player["game_code_verified"] = (
                        is_participating_player(room, roster_player)
                        or not room.game_code_enabled
                        or provided_code == room.game_code
                    )
                    room.players[visitor_id] = roster_player
                elif visitor_id in room.players:
                    room.players[visitor_id]["ws"] = websocket
                    room.players[visitor_id]["name"] = name
                    room.players[visitor_id]["student_number"] = student_number
                    room.players[visitor_id]["game_code_verified"] = (
                        is_participating_player(room, room.players[visitor_id])
                        or not room.game_code_enabled
                        or provided_code == room.game_code
                    )
                else:
                    room.players[visitor_id] = {
                        "name": name,
                        "student_number": student_number,
                        "score": 0,
                        "streak": 0,
                        "answers": [],
                        "ws": websocket,
                        "game_code_verified": (not room.game_code_enabled or provided_code == room.game_code)
                    }
                # Remember them for the rest of the session, whatever happens
                # to their live connection from here on.
                room.session_roster[student_key] = room.players[visitor_id]
                await websocket.send_text(json.dumps(build_joined_payload(room, visitor_id)))
                await push_room_update(room)
                await sync_answer_count(room)

            elif action == "verify_game_code":
                if role != "player" or room is None or visitor_id not in room.players:
                    continue
                if not room.game_code_enabled:
                    await websocket.send_text(json.dumps(build_joined_payload(room, visitor_id)))
                    continue
                provided_code = (msg.get("gameCode") or "").strip()
                if provided_code != room.game_code:
                    await websocket.send_text(json.dumps({
                        "type": "error_game_code",
                        "message": "Enter the 4-digit code shown on the lecturer screen to continue."
                    }))
                    continue
                room.players[visitor_id]["game_code_verified"] = True
                await websocket.send_text(json.dumps(build_joined_payload(room, visitor_id)))
                await push_room_update(room)
                await sync_answer_count(room)

            elif action == "player_leave":
                if role != "player" or room is None:
                    continue
                room.players.pop(visitor_id, None)
                await websocket.send_text(json.dumps({"type": "left"}))
                await push_room_update(room)
                await sync_answer_count(room)
                await maybe_finish_question_early(room)
                await websocket.close()
                break

            elif action == "answer":
                if role != "player" or room is None or room.phase != "question":
                    continue
                if not is_participating_player(room, room.players.get(visitor_id)):
                    await websocket.send_text(json.dumps({
                        "type": "error_game_code",
                        "message": "Enter the 4-digit code shown on the lecturer screen to continue."
                    }))
                    continue
                if visitor_id in room.answers_this_round:
                    continue
                # Validate the choice. It used to be taken as-is, so a client
                # could submit a string, a list or an out-of-range index and it
                # would be stored and later written into the results workbook.
                raw_choice = msg.get("choice", -1)
                if isinstance(raw_choice, bool) or not isinstance(raw_choice, int):
                    # Not "error": the player client treats that as a join
                    # failure and would throw the student back to the join screen.
                    await websocket.send_text(json.dumps({"type": "invalid_answer", "message": "That answer was not understood."}))
                    continue
                choice = raw_choice
                if not 0 <= choice < len(room.questions[room.current_q]["options"]):
                    # Not "error": the player client treats that as a join
                    # failure and would throw the student back to the join screen.
                    await websocket.send_text(json.dumps({"type": "invalid_answer", "message": "That answer was not understood."}))
                    continue
                answer_time = time.time() - room.question_start_time
                room.answers_this_round[visitor_id] = {"choice": choice, "time": answer_time}

                q = room.questions[room.current_q]
                is_correct = choice == q["correct"]
                points = 0
                if is_correct:
                    # Against the resolved limit, so a 90 second question still awards
                    # full marks for a fast answer.
                    time_fraction = min(answer_time / room.time_limit_for(), 1.0)
                    points = round(MAX_POINTS - (MAX_POINTS - MIN_POINTS) * time_fraction)
                    room.players[visitor_id]["streak"] += 1
                    if room.players[visitor_id]["streak"] >= 3:
                        points = round(points * 1.2)
                else:
                    room.players[visitor_id]["streak"] = 0
                room.players[visitor_id]["score"] += points
                room.players[visitor_id]["answers"].append({
                    "q": room.current_q,
                    "choice": choice,
                    "correct": is_correct,
                    "points": points,
                    "time": answer_time
                })
                await websocket.send_text(json.dumps({
                    "type": "answer_result",
                    "correct": is_correct,
                    "points": points,
                    "totalScore": room.players[visitor_id]["score"],
                    "streak": room.players[visitor_id]["streak"],
                    "correctAnswer": q["correct"],
                    "explanation": q["explanation"]
                }))
                await sync_answer_count(room)
                await maybe_finish_question_early(room)

    except WebSocketDisconnect:
        if role == "host" and room:
            room.host_ws = None
        elif role == "player" and room and visitor_id in room.players:
            # Only clear ws if it's still the current connection — a reconnect may
            # have already replaced it, and we must not overwrite the new socket.
            if room.players[visitor_id].get("ws") is websocket:
                room.players[visitor_id]["ws"] = None
                await push_room_update(room)
                await sync_answer_count(room)
                await maybe_finish_question_early(room)
    except Exception:
        if role == "host" and room:
            room.host_ws = None
        elif role == "player" and room and visitor_id in room.players:
            if room.players[visitor_id].get("ws") is websocket:
                room.players[visitor_id]["ws"] = None
                await push_room_update(room)
                await sync_answer_count(room)
                await maybe_finish_question_early(room)


async def run_game_start(room: GameRoom, *, shuffle: bool, use_code: bool) -> None:
    """Start sequence, run as a background task so the host socket stays live."""
    if use_code:
        room.game_code = generate_game_code()
        room.game_code_enabled = True
        for player in room.players.values():
            player["game_code_verified"] = False
    else:
        room.game_code = ""
        room.game_code_enabled = False
        for player in room.players.values():
            player["game_code_verified"] = True

    if room.game_code_enabled:
        await send_to_host(room, {
            "type": "game_code_display",
            "code": room.game_code,
            "countdown": GAME_CODE_COUNTDOWN_SECONDS
        })
        await broadcast_to_players(room, {
            "type": "game_code_required",
            "countdown": GAME_CODE_COUNTDOWN_SECONDS
        })
        await asyncio.sleep(GAME_CODE_COUNTDOWN_SECONDS)
        if room.phase != "lobby":
            return       # cancelled or reset while the code was showing

    room.paused = False
    if shuffle:
        random.shuffle(room.questions)
    room.phase = "get_ready"
    room.current_q = 0
    for player in list(room.players.values()) + list(room.session_roster.values()):
        player["score"] = 0
        player["streak"] = 0
        player["answers"] = []
    await broadcast_to_players(room, {"type": "get_ready", "qNum": 1, "totalQ": room.total_q}, participant_only=True)
    await send_to_host(room, {"type": "get_ready", "qNum": 1, "totalQ": room.total_q})
    await asyncio.sleep(GET_READY_SECONDS)
    if room.phase == "get_ready" and room.current_q == 0:
        await send_question(room)


async def send_question(room: GameRoom) -> None:
    q = room.questions[room.current_q]
    q_index = room.current_q
    room.phase = "question"
    server_ts = time.time()
    room.question_start_time = server_ts
    room.question_elapsed = 0.0
    room.question_time_limit = room.time_limit_for(q_index)
    room.answers_this_round = {}
    limit = room.question_time_limit
    await broadcast_to_players(room, {
        "type": "question",
        "qNum": room.current_q + 1,
        "totalQ": room.total_q,
        "question": q["q"],
        "options": q["options"],
        "timeLimit": limit,
        "serverTimestamp": server_ts
    }, participant_only=True)
    await send_to_host(room, {
        "type": "question",
        "qNum": room.current_q + 1,
        "totalQ": room.total_q,
        "question": q["q"],
        "options": q["options"],
        "correctAnswer": q["correct"],
        "timeLimit": limit,
        "serverTimestamp": server_ts
    })
    await sync_answer_count(room)

    async def _timer():
        # Count elapsed time in 0.5s ticks, freezing while room.paused is True
        elapsed = 0.0
        while elapsed < room.question_time_limit:
            await asyncio.sleep(0.5)
            if room.phase != "question" or room.current_q != q_index:
                return  # Question already advanced (e.g. all answered early)
            if not room.paused:
                elapsed += 0.5
                room.question_elapsed = elapsed
        async with room.advance_lock:
            if room.phase != "question" or room.current_q != q_index:
                return
            mark_unanswered_players(room)
            room.phase = "reveal"
        await auto_reveal(room)

    room.question_timer_task = asyncio.create_task(_timer())


async def auto_reveal(room: GameRoom) -> None:
    q_index = room.current_q
    q = room.questions[q_index]
    room.phase = "reveal"
    lb = get_leaderboard(room, participant_only=True)
    for vid, player in room.players.items():
        if not is_participating_player(room, player):
            continue
        if vid not in room.answers_this_round:
            ws = player.get("ws")
            if ws:
                try:
                    await ws.send_text(json.dumps({
                        "type": "answer_result",
                        "correct": False,
                        "points": 0,
                        "totalScore": player["score"],
                        "correctAnswer": q["correct"],
                        "explanation": q["explanation"],
                        "timedOut": True
                    }))
                except Exception:
                    pass
    await broadcast_to_players(room, {"type": "leaderboard", "leaderboard": lb}, participant_only=True)
    await send_to_host(room, {
        "type": "reveal",
        "correctAnswer": q["correct"],
        "explanation": q["explanation"],
        "leaderboard": lb,
        "isLast": room.current_q >= room.total_q - 1,
        "revealSeconds": REVEAL_SECONDS,
        "options": q.get("options", []),
        "question": q.get("q", ""),
        "distribution": build_answer_distribution(room),
        "answered": len(room.answers_this_round),
    })
    waited = 0.0
    tick = min(0.5, max(0.05, REVEAL_SECONDS / 4))
    while waited < REVEAL_SECONDS:
        await asyncio.sleep(tick)
        if room.paused:
            continue
        waited += tick
    if room.phase == "reveal" and room.current_q == q_index:
        await advance_to_next(room, from_question=q_index)


async def advance_to_next(room: GameRoom, *, from_question: int | None = None) -> None:
    """Move on to the next question.

    Guarded so that a double click on "Next Question", or a click racing the
    auto-reveal timer, cannot increment current_q twice and skip a question.
    `from_question` is the index the caller believed was current; if the room
    has already moved past it, this call is a duplicate and does nothing.
    """
    async with room.advance_lock:
        if room.phase in ("get_ready", "final"):
            return                      # an advance is already under way
        if from_question is not None and room.current_q != from_question:
            return                      # someone else advanced first
        room.current_q += 1
        finished = room.current_q >= room.total_q
        room.phase = "final" if finished else "get_ready"
        next_index = room.current_q

    if finished:
        await force_end_game(room)
        return

    await broadcast_to_players(room, {"type": "get_ready", "qNum": next_index + 1, "totalQ": room.total_q}, participant_only=True)
    await send_to_host(room, {"type": "get_ready", "qNum": next_index + 1, "totalQ": room.total_q})
    await asyncio.sleep(GET_READY_SECONDS)
    if room.phase == "get_ready" and room.current_q == next_index:
        await send_question(room)


if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", "8000"))
    uvicorn.run(app, host="0.0.0.0", port=port)
